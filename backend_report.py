# from flask import Blueprint, Flask, jsonify, request
# from flask_cors import CORS
# from config import Config
# from sqlalchemy import or_
# from models import Document, db, VehicleTypeMaster, VehicleTypeDetails,DockInOutDetails,TripStatus, MasterVehicle, MasterDriver
# from datetime import datetime

# app = Flask(__name__)
# CORS(app)
# app.config.from_object(Config)
# db.init_app(app)



# # ------------------------------------------------------------
# # 1️⃣ Fetch only master table data
# # ------------------------------------------------------------
# @app.route("/get_vehicle_types", methods=["GET"])
# def get_vehicle_types():
#     """Returns only VehicleTypeMaster data"""
#     try:
#         masters = VehicleTypeMaster.query.all()
#         result = [m.to_dict() for m in masters]
#         return jsonify({
#             "status": "success",
#             "count": len(result),
#             "data": result
#         }), 200
#     except Exception as e:
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500


# # ------------------------------------------------------------
# # 2️⃣ Fetch master + details (joined)
# # ------------------------------------------------------------
# @app.route("/get_vehicle_type_details", methods=["GET"])
# def get_vehicle_type_details():
#     """Returns full master-detail vehicle type data"""
#     try:
#         results = (
#             db.session.query(
#                 VehicleTypeMaster.id,
#                 VehicleTypeMaster.vehicle_type,
#                 VehicleTypeMaster.plant_location,
#                 VehicleTypeMaster.vehicle_type_created_by,
#                 VehicleTypeMaster.vehicle_type_created_time,
#                 VehicleTypeDetails.id.label("detail_id"),
#                 VehicleTypeDetails.status,
#                 VehicleTypeDetails.unload,
#                 VehicleTypeDetails.category,
#                 VehicleTypeDetails.standard_time_minutes,
#                 VehicleTypeDetails.standard_weight_kgs,
#                 VehicleTypeDetails.created_by,
#                 VehicleTypeDetails.created_time,
#                 VehicleTypeDetails.modified_by,
#                 VehicleTypeDetails.last_modified_time
#             )
#             .join(VehicleTypeDetails, VehicleTypeMaster.id == VehicleTypeDetails.vehicle_type_id)
#             .order_by(VehicleTypeMaster.vehicle_type)
#             .all()
#         )

#         data = [
#             {
#                 "vehicle_type_master_id": r.id,
#                 "vehicle_type": r.vehicle_type,
#                 "plant_location": r.plant_location,
#                 "vehicle_type_created_by": r.vehicle_type_created_by,
#                 "vehicle_type_created_time": r.vehicle_type_created_time,
#                 "detail_id": r.detail_id,
#                 "status": r.status,
#                 "unload": r.unload,
#                 "category": r.category,
#                 "standard_time_minutes": r.standard_time_minutes,
#                 "standard_weight_kgs": r.standard_weight_kgs,
#                 "created_by": r.created_by,
#                 "created_time": r.created_time,
#                 "modified_by": r.modified_by,
#                 "last_modified_time": r.last_modified_time
#             }
#             for r in results
#         ]

#         return jsonify({
#             "status": "success",
#             "count": len(data),
#             "data": data
#         }), 200

#     except Exception as e:
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500


# # ------------------------------------------------------------
# # 3️⃣ Create new vehicle type (no duplicates)
# # ------------------------------------------------------------
# @app.route("/create_vehicle_type", methods=["POST"])
# def create_vehicle_type():
#     """Add new vehicle type — no duplicates allowed"""
#     try:
#         data = request.get_json()
#         vehicle_type = data.get("vehicle_type", "").strip()
#         plant_location = data.get("plant_location", "KAGAL")
#         created_by = data.get("created_by", "admin")

#         if not vehicle_type:
#             return jsonify({"status": "error", "message": "Vehicle type is required"}), 400

#         # Check duplicate (case-insensitive)
#         existing = VehicleTypeMaster.query.filter(
#             db.func.lower(VehicleTypeMaster.vehicle_type) == vehicle_type.lower()
#         ).first()
#         if existing:
#             return jsonify({"status": "error", "message": "Vehicle type already exists"}), 409

#         new_type = VehicleTypeMaster(
#             vehicle_type=vehicle_type,
#             plant_location=plant_location,
#             vehicle_type_created_by=created_by,
#             vehicle_type_created_time=datetime.now()
#         )

#         db.session.add(new_type)
#         db.session.commit()

#         return jsonify({
#             "status": "success",
#             "message": "Vehicle type created successfully"
#         }), 201

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"status": "error", "message": str(e)}), 500


# # ------------------------------------------------------------
# # 4️⃣ Save / Update vehicle type detail
# # ------------------------------------------------------------
# @app.route("/save_vehicle_type_details", methods=["POST"])
# def save_vehicle_type_details():
#     """Insert or update details per dropdown combination"""
#     try:
#         data = request.get_json()
#         vehicle_type_id = data.get("vehicle_type_id")
#         status = data.get("status")
#         unload = data.get("unload")
#         category = data.get("category")
#         standard_time_minutes = data.get("standard_time_minutes")
#         standard_weight_kgs = data.get("standard_weight_kgs")
#         created_by = data.get("created_by", "admin")

#         if not all([vehicle_type_id, status, unload, category]):
#             return jsonify({"status": "error", "message": "All dropdowns are required"}), 400

#         existing = VehicleTypeDetails.query.filter_by(
#             vehicle_type_id=vehicle_type_id,
#             status=status,
#             unload=unload,
#             category=category
#         ).first()

#         if existing:
#             existing.standard_time_minutes = standard_time_minutes
#             existing.standard_weight_kgs = standard_weight_kgs
#             existing.modified_by = created_by
#             existing.last_modified_time = datetime.now()
#         else:
#             new_detail = VehicleTypeDetails(
#                 vehicle_type_id=vehicle_type_id,
#                 status=status,
#                 unload=unload,
#                 category=category,
#                 standard_time_minutes=standard_time_minutes,
#                 standard_weight_kgs=standard_weight_kgs,
#                 created_by=created_by,
#                 created_time=datetime.now()
#             )
#             db.session.add(new_detail)

#         db.session.commit()
#         return jsonify({"status": "success", "message": "Saved successfully"}), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"status": "error", "message": str(e)}), 500
    

# @app.route("/vehicle_type/<int:vehicle_type_id>", methods=["DELETE"])
# def delete_vehicle_type(vehicle_type_id):
#     try:
#         # Get optional query params for deleting specific detail
#         status = request.args.get("status")
#         unload = request.args.get("unload")
#         category = request.args.get("category")

#         # Delete specific detail if all params are provided
#         if status and unload and category:
#             detail = VehicleTypeDetails.query.filter_by(
#                 vehicle_type_id=vehicle_type_id,
#                 status=status,
#                 unload=unload,
#                 category=category
#             ).first()
#             if not detail:
#                 return jsonify({"message": "No matching vehicle type detail found"}), 404

#             db.session.delete(detail)
#             db.session.commit()
#             return jsonify({"message": "Vehicle type detail deleted successfully"}), 200

#         # Else delete master and all related details
#         master = VehicleTypeMaster.query.filter_by(id=vehicle_type_id).first()
#         if not master:
#             return jsonify({"message": "Vehicle type master not found"}), 404

#         db.session.delete(master)
#         db.session.commit()
#         return jsonify({"message": "Vehicle type and all related details deleted successfully"}), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"message": f"Error deleting vehicle type: {str(e)}"}), 500
    
# from flask import Blueprint, Flask, jsonify, request
# from flask_cors import CORS
# from config import Config
# from sqlalchemy import or_
# from models import db, VehicleTypeMaster, VehicleTypeDetails, MasterVehicle, MasterDriver
# from datetime import datetime

# app = Flask(__name__)
# CORS(app)
# app.config.from_object(Config)
# db.init_app(app)



# # ------------------------------------------------------------
# # 1️⃣ Fetch only master table data
# # ------------------------------------------------------------
# @app.route("/get_vehicle_types", methods=["GET"])
# def get_vehicle_types():
#     """Returns only VehicleTypeMaster data"""
#     try:
#         masters = VehicleTypeMaster.query.all()
#         result = [m.to_dict() for m in masters]
#         return jsonify({
#             "status": "success",
#             "count": len(result),
#             "data": result
#         }), 200
#     except Exception as e:
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500


# # ------------------------------------------------------------
# # 2️⃣ Fetch master + details (joined)
# # ------------------------------------------------------------
# @app.route("/get_vehicle_type_details", methods=["GET"])
# def get_vehicle_type_details():
#     """Returns full master-detail vehicle type data"""
#     try:
#         results = (
#             db.session.query(
#                 VehicleTypeMaster.id,
#                 VehicleTypeMaster.vehicle_type,
#                 VehicleTypeMaster.plant_location,
#                 VehicleTypeMaster.vehicle_type_created_by,
#                 VehicleTypeMaster.vehicle_type_created_time,
#                 VehicleTypeDetails.id.label("detail_id"),
#                 VehicleTypeDetails.status,
#                 VehicleTypeDetails.unload,
#                 VehicleTypeDetails.category,
#                 VehicleTypeDetails.standard_time_minutes,
#                 VehicleTypeDetails.standard_weight_kgs,
#                 VehicleTypeDetails.created_by,
#                 VehicleTypeDetails.created_time,
#                 VehicleTypeDetails.modified_by,
#                 VehicleTypeDetails.last_modified_time
#             )
#             .join(VehicleTypeDetails, VehicleTypeMaster.id == VehicleTypeDetails.vehicle_type_id)
#             .order_by(VehicleTypeMaster.vehicle_type)
#             .all()
#         )

#         data = [
#             {
#                 "vehicle_type_master_id": r.id,
#                 "vehicle_type": r.vehicle_type,
#                 "plant_location": r.plant_location,
#                 "vehicle_type_created_by": r.vehicle_type_created_by,
#                 "vehicle_type_created_time": r.vehicle_type_created_time,
#                 "detail_id": r.detail_id,
#                 "status": r.status,
#                 "unload": r.unload,
#                 "category": r.category,
#                 "standard_time_minutes": r.standard_time_minutes,
#                 "standard_weight_kgs": r.standard_weight_kgs,
#                 "created_by": r.created_by,
#                 "created_time": r.created_time,
#                 "modified_by": r.modified_by,
#                 "last_modified_time": r.last_modified_time
#             }
#             for r in results
#         ]

#         return jsonify({
#             "status": "success",
#             "count": len(data),
#             "data": data
#         }), 200

#     except Exception as e:
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500


# # ------------------------------------------------------------
# # 3️⃣ Create new vehicle type (no duplicates)
# # ------------------------------------------------------------
# @app.route("/create_vehicle_type", methods=["POST"])
# def create_vehicle_type():
#     """Add new vehicle type — no duplicates allowed"""
#     try:
#         data = request.get_json()
#         vehicle_type = data.get("vehicle_type", "").strip()
#         plant_location = data.get("plant_location", "KAGAL")
#         created_by = data.get("created_by", "admin")

#         if not vehicle_type:
#             return jsonify({"status": "error", "message": "Vehicle type is required"}), 400

#         # Check duplicate (case-insensitive)
#         existing = VehicleTypeMaster.query.filter(
#             db.func.lower(VehicleTypeMaster.vehicle_type) == vehicle_type.lower()
#         ).first()
#         if existing:
#             return jsonify({"status": "error", "message": "Vehicle type already exists"}), 409

#         new_type = VehicleTypeMaster(
#             vehicle_type=vehicle_type,
#             plant_location=plant_location,
#             vehicle_type_created_by=created_by,
#             vehicle_type_created_time=datetime.now()
#         )

#         db.session.add(new_type)
#         db.session.commit()

#         return jsonify({
#             "status": "success",
#             "message": "Vehicle type created successfully"
#         }), 201

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"status": "error", "message": str(e)}), 500


# # ------------------------------------------------------------
# # 4️⃣ Save / Update vehicle type detail
# # ------------------------------------------------------------
# @app.route("/save_vehicle_type_details", methods=["POST"])
# def save_vehicle_type_details():
#     """Insert or update details per dropdown combination"""
#     try:
#         data = request.get_json()
#         vehicle_type_id = data.get("vehicle_type_id")
#         status = data.get("status")
#         unload = data.get("unload")
#         category = data.get("category")
#         standard_time_minutes = data.get("standard_time_minutes")
#         standard_weight_kgs = data.get("standard_weight_kgs")
#         created_by = data.get("created_by", "admin")

#         if not all([vehicle_type_id, status, unload, category]):
#             return jsonify({"status": "error", "message": "All dropdowns are required"}), 400

#         existing = VehicleTypeDetails.query.filter_by(
#             vehicle_type_id=vehicle_type_id,
#             status=status,
#             unload=unload,
#             category=category
#         ).first()

#         if existing:
#             existing.standard_time_minutes = standard_time_minutes
#             existing.standard_weight_kgs = standard_weight_kgs
#             existing.modified_by = created_by
#             existing.last_modified_time = datetime.now()
#         else:
#             new_detail = VehicleTypeDetails(
#                 vehicle_type_id=vehicle_type_id,
#                 status=status,
#                 unload=unload,
#                 category=category,
#                 standard_time_minutes=standard_time_minutes,
#                 standard_weight_kgs=standard_weight_kgs,
#                 created_by=created_by,
#                 created_time=datetime.now()
#             )
#             db.session.add(new_detail)

#         db.session.commit()
#         return jsonify({"status": "success", "message": "Saved successfully"}), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"status": "error", "message": str(e)}), 500
    

# @app.route("/vehicle_type/<int:vehicle_type_id>", methods=["DELETE"])
# def delete_vehicle_type(vehicle_type_id):
#     try:
#         # Get optional query params for deleting specific detail
#         status = request.args.get("status")
#         unload = request.args.get("unload")
#         category = request.args.get("category")

#         # Delete specific detail if all params are provided
#         if status and unload and category:
#             detail = VehicleTypeDetails.query.filter_by(
#                 vehicle_type_id=vehicle_type_id,
#                 status=status,
#                 unload=unload,
#                 category=category
#             ).first()
#             if not detail:
#                 return jsonify({"message": "No matching vehicle type detail found"}), 404

#             db.session.delete(detail)
#             db.session.commit()
#             return jsonify({"message": "Vehicle type detail deleted successfully"}), 200

#         # Else delete master and all related details
#         master = VehicleTypeMaster.query.filter_by(id=vehicle_type_id).first()
#         if not master:
#             return jsonify({"message": "Vehicle type master not found"}), 404

#         db.session.delete(master)
#         db.session.commit()
#         return jsonify({"message": "Vehicle type and all related details deleted successfully"}), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"message": f"Error deleting vehicle type: {str(e)}"}), 500
    
# vehicle_bp = Blueprint('vehicle_bp', __name__)

  
# @app.route('/get_vehicles', methods=['GET'])
# def get_vehicles():
#     try:
#         # --- 1️⃣ Get query params ---
#         search = request.args.get('search', '', type=str)
#         page = request.args.get('page', 1, type=int)
#         limit = request.args.get('limit', 20, type=int)

#         # --- 2️⃣ Base query ---
#         query = MasterVehicle.query

#         # --- 3️⃣ Apply search filter if provided ---
#         if search:
#             query = query.filter(MasterVehicle.vehicle_number.ilike(f"%{search}%"))

#         # --- 4️⃣ Select only required fields AFTER filtering ---
#         query = query.with_entities(MasterVehicle.id, MasterVehicle.vehicle_number)

#         # --- 5️⃣ Pagination ---
#         pagination = query.order_by(MasterVehicle.id.desc()).paginate(page=page, per_page=limit, error_out=False)

#         # --- 6️⃣ Prepare response ---
#         data = [{"id": v.id, "vehicle_number": v.vehicle_number} for v in pagination.items]

#         return jsonify({
#             "status": "success",
#             "page": page,
#             "total_pages": pagination.pages,
#             "total_items": pagination.total,
#             "vehicles": data
#         }), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500

    

# @app.route('/get_vehicles/<int:vehicle_id>', methods=['GET'])
# def get_vehicle(vehicle_id):
#     try:
#         # Fetch vehicle details
#         vehicle = MasterVehicle.query.get(vehicle_id)
#         if not vehicle:
#             return jsonify({'message': 'Vehicle not found'}), 404

#         # Count number of times vehicle entered plant
#         entry_count = Document.query.filter_by(vehicle_number=vehicle.vehicle_number).count()

#         data = {
#             'id': vehicle.id,
#             'vehicle_number': vehicle.vehicle_number,
#             'vehicle_type': vehicle.vehicle_type,
#             'pravesh_remarks': vehicle.vehicle_doc_remark,
#             'puc_validity': vehicle.puc_validity,
#             'insurance_validity': vehicle.insurance_validity,
#             'green_channel': vehicle.green_channel,
#             'vehicle_banned': vehicle.vehicle_banned,
#             'vehicle_banned_remarks': vehicle.vehicle_banned_remarks,
#             'entry_count': entry_count  # 👈 new field added
#         }

#         return jsonify(data), 200

#     except Exception as e:
#         return jsonify({'error': str(e)}), 500
    
# from sqlalchemy.exc import SQLAlchemyError


# @app.route("/update_vehicle_status", methods=["PUT"])
# def update_vehicle_status():
#     try:
#         data = request.get_json()
#         vehicle_id = data.get("id")
#         green_channel = data.get("green_channel")
#         vehicle_banned = data.get("vehicle_banned")
#         vehicle_banned_remarks = data.get("vehicle_banned_remarks")

#         if not vehicle_id:
#             return jsonify({"error": "Vehicle ID is required"}), 400

#         vehicle = MasterVehicle.query.get(vehicle_id)
#         if not vehicle:
#             return jsonify({"error": "Vehicle not found"}), 404

#         # ✅ Update only the provided fields
#         if green_channel is not None:
#             vehicle.green_channel = green_channel
#         if vehicle_banned is not None:
#             vehicle.vehicle_banned = vehicle_banned
#         if vehicle_banned_remarks is not None:
#             vehicle.vehicle_banned_remarks = vehicle_banned_remarks

#         db.session.commit()

#         return jsonify({
#             "message": "Vehicle status updated successfully",
#             "updated_data": {
#                 "id": vehicle.id,
#                 "green_channel": vehicle.green_channel,
#                 "vehicle_banned": vehicle.vehicle_banned,
#                 "vehicle_banned_remarks": vehicle.vehicle_banned_remarks
#             }
#         }), 200

#     except Exception as e:
#         db.session.rollback()
#         print("Error:", str(e))
#         return jsonify({"error": str(e)}), 500
    

# @app.route('/get_drivers', methods=['GET'])
# def fetch_drivers():
#     try:
#         search = request.args.get('search', '', type=str)
#         page = request.args.get('page', 1, type=int)
#         limit = request.args.get('limit', 20, type=int)

#         query = MasterDriver.query

#         if search:
#             query = query.filter(
#                 db.or_(
#                     MasterDriver.driver_license_number.ilike(f"%{search}%"),
#                     MasterDriver.driver_mobile_number.ilike(f"%{search}%"),
#                     MasterDriver.driver_vehicle_number.ilike(f"%{search}%")
#                 )
#             )

#         # Fetch extracted_text_driver also
#         query = query.with_entities(
#             MasterDriver.id,
#             MasterDriver.driver_license_number,
#             MasterDriver.driver_mobile_number,
#             MasterDriver.driver_vehicle_number,
#             MasterDriver.extracted_text_driver
#         )

#         pagination = query.order_by(MasterDriver.id.desc()).paginate(
#             page=page, per_page=limit, error_out=False
#         )

#         data = []
#         for d in pagination.items:
#             # Parse extracted_text_driver → name
#             try:
#                 extracted = (
#                     d.extracted_text_driver
#                     if isinstance(d.extracted_text_driver, dict)
#                     else json.loads(d.extracted_text_driver or "{}")
#                 )
#                 name = extracted.get("name", "")
#             except:
#                 name = ""

#             data.append({
#                 "id": d.id,
#                 "driver_license_number": d.driver_license_number,
#                 "driver_mobile_number": d.driver_mobile_number,
#                 "driver_vehicle_number": d.driver_vehicle_number,
#                 "driver_name": name    # <-- added here
#             })

#         return jsonify({
#             "status": "success",
#             "page": page,
#             "total_pages": pagination.pages,
#             "total_items": pagination.total,
#             "drivers": data
#         }), 200

#     except Exception as e:
#         db.session.rollback()
#         return jsonify({"status": "error", "message": str(e)}), 500



    

    


# import json

# @app.route('/get_driver/<int:driver_id>', methods=['GET'])
# def fetch_driver_details(driver_id):
#     try:
#         driver = MasterDriver.query.get(driver_id)
#         if not driver:
#             return jsonify({'message': 'Driver not found'}), 404

#         # Parse extracted text safely
#         extracted_data = {}
#         if driver.extracted_text_driver:
#             try:
#                 extracted_data = (
#                     driver.extracted_text_driver
#                     if isinstance(driver.extracted_text_driver, dict)
#                     else json.loads(driver.extracted_text_driver)
#                 )
#             except:
#                 extracted_data = {}

#         driver_name = extracted_data.get("name", "")

#         trip_count = Document.query.filter_by(
#             driver_license_number=driver.driver_license_number
#         ).count()

#         data = {
#             'id': driver.id,
#             'driver_mobile_number': driver.driver_mobile_number,
#             'driver_license_number': driver.driver_license_number,
#             'driver_vehicle_number': driver.driver_vehicle_number,
#             'license_validity': driver.license_validity,
#             'driver_name': driver_name,
#             'created_at': driver.created_at,
#             'driver_banned': driver.driver_banned,
#             'driver_banned_remarks': driver.driver_banned_remarks,
#             'trip_count': trip_count
#         }

#         return jsonify(data), 200

#     except Exception as e:
#         return jsonify({'error': str(e)}), 500
    

# from datetime import datetime, timedelta
# import pandas as pd
# import os



# @app.route('/get_report', methods=['GET'])
# def get_report():
#     try:
#         start_date = request.args.get('startDate')
#         end_date = request.args.get('endDate')

#         start = datetime.strptime(start_date, "%Y-%m-%d")
#         end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

#         documents = Document.query.filter(
#             Document.timestamp >= start,
#             Document.timestamp <= end
#         ).all()

#         results = []

#         for doc in documents:

#             # Fetch standard weight
#             vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
#             standard_weight = vt.standard_weight_kgs if vt else None

#             # Fetch Dock-In & Dock-Out users
#             dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
#             user_dock_in = list({d.user_dockin for d in dock_details if d.user_dockin})
#             user_dock_out = list({d.user_dockout for d in dock_details if d.user_dockout})

#             # Handle "Other Transporter"
#             transporter = doc.transporter_name
#             if transporter == "Other":
#                 transporter = doc.other_transporter_name or transporter

#             results.append({
#                 "Trip ID": doc.Trip_id,
#                 "Pravesh Entry Date": doc.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
#                 "Vehicle No": doc.vehicle_number,
#                 "Total Invoices": doc.total_invoices,
#                 "Loading / Unloading": doc.loading_or_unloading,
#                 "Transporter": transporter,
#                 "Vehicle Type": doc.vehicle_type,
#                 "Standard Weight": standard_weight,
#                 "User-Dock-In": user_dock_in,
#                 "User-Dock-Out": user_dock_out
#             })

#         return jsonify({"status": "success", "data": results})

#     except Exception as e:
#         return jsonify({"status": "error", "message": str(e)}), 500


# @app.route('/download_report', methods=['GET'])
# def download_report():
#     try:
#         start_date = request.args.get('startDate')
#         end_date = request.args.get('endDate')

#         start = datetime.strptime(start_date, "%Y-%m-%d")
#         end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

#         documents = Document.query.filter(
#             Document.timestamp >= start,
#             Document.timestamp <= end
#         ).all()

#         from openpyxl import Workbook
#         from openpyxl.styles import Font
#         import io

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Report"

#         headers = [
#             "Trip ID",
#             "Pravesh Entry Date",
#             "Vehicle No",
#             "Total Invoices",
#             "Loading / Unloading",
#             "Transporter",
#             "Vehicle Type",
#             "Standard Weight (kgs)",
#             "User-Dock-In",
#             "User-Dock-Out"
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for doc in documents:
#             vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
#             standard_weight = vt.standard_weight_kgs if vt else None

#             dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
#             user_dock_in = ", ".join(list({d.user_dockin for d in dock_details if d.user_dockin}))
#             user_dock_out = ", ".join(list({d.user_dockout for d in dock_details if d.user_dockout}))

#             # Handle "Other Transporter"
#             transporter = doc.transporter_name
#             if transporter == "Other":
#                 transporter = doc.other_transporter_name or transporter

#             ws.append([
#                 doc.Trip_id,
#                 doc.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
#                 doc.vehicle_number,
#                 doc.total_invoices,
#                 doc.loading_or_unloading,
#                 transporter,
#                 doc.vehicle_type,
#                 standard_weight,
#                 user_dock_in,
#                 user_dock_out
#             ])

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=f"Report_{start_date}_to_{end_date}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         return jsonify({"status": "error", "message": str(e)}), 500
    


# @app.route('/download_report_vehicle', methods=['GET'])
# def download_report_vehicle():
#     try:
#         start_date = request.args.get('startDate')
#         end_date = request.args.get('endDate')
#         vehicle_number = request.args.get('vehicleNumber')

#         if not vehicle_number or vehicle_number.strip() == "":
#             return jsonify({"status": "error", "message": "vehicleNumber is required"}), 400

#         vehicle_number = vehicle_number.strip()  # Remove spaces

#         # Parse dates
#         start = datetime.strptime(start_date, "%Y-%m-%d")
#         end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

#         # Filter documents by date range and vehicle_number (case-insensitive, partial match)
#         documents = Document.query.filter(
#             Document.timestamp >= start,
#             Document.timestamp < end,
#             Document.vehicle_number.ilike(f"%{vehicle_number}%")
#         ).all()

#         if not documents:
#             return jsonify({"status": "error", "message": "No records found for the given vehicle and date range"}), 404

#         # Create Excel workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = f"Report_{vehicle_number}"

#         headers = [
#             "Trip ID",
#             "Pravesh Entry Date",
#             "Vehicle No",
#             "Total Invoices",
#             "Loading / Unloading",
#             "Transporter",
#             "Vehicle Type",
#             "Standard Weight (kgs)",
#             "User-Dock-In",
#             "User-Dock-Out"
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for doc in documents:
#             try:
#                 vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
#                 standard_weight = vt.standard_weight_kgs if vt else None

#                 dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
#                 user_dock_in = ", ".join(sorted({d.user_dockin for d in dock_details if d.user_dockin}))
#                 user_dock_out = ", ".join(sorted({d.user_dockout for d in dock_details if d.user_dockout}))

#                 transporter = doc.transporter_name
#                 if transporter == "Other":
#                     transporter = doc.other_transporter_name or transporter

#                 ws.append([
#                     doc.Trip_id or "",
#                     doc.timestamp.strftime('%Y-%m-%d %H:%M:%S') if doc.timestamp else "",
#                     doc.vehicle_number or "",
#                     doc.total_invoices or 0,
#                     doc.loading_or_unloading or "",
#                     transporter,
#                     doc.vehicle_type or "",
#                     standard_weight or "",
#                     user_dock_in,
#                     user_dock_out
#                 ])
#             except Exception as row_err:
#                 print(f"Skipping row due to error: {row_err}")
#                 continue

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=f"Report_{vehicle_number}_{start_date}_to_{end_date}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         print(f"Download vehicle report error: {e}")
#         return jsonify({"status": "error", "message": str(e)}), 500
    

# from models import TripStatus, DockInOutDetails, Document  # ensure import

# from models import TripStatus, DockInOutDetails, Document  # make sure this import exists
# import json  # make sure this is imported
    
# @app.route("/api/report/date-range", methods=["GET"])
# def trip_report_date_range():
#     try:
#         # ---- GET query parameters ----
#         from_date_str = request.args.get("from_date")
#         to_date_str = request.args.get("to_date")
#         page = int(request.args.get("page", 1))
#         limit = int(request.args.get("limit", 50))

#         if not from_date_str or not to_date_str:
#             return jsonify({"status": "error", "message": "from_date & to_date required YYYY-MM-DD"}), 400

#         # ---- Parse only dates ----
#         from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
#         to_date = datetime.strptime(to_date_str, "%Y-%m-%d")

#         # ---- Convert to full datetime range ----
#         from_date = datetime.combine(from_date, datetime.min.time())   # 00:00:00
#         to_date = datetime.combine(to_date, datetime.max.time())       # 23:59:59

#         # ---- Filter TripStatus for GATE ENTRY ----
#         filtered_trips = (
#             db.session.query(TripStatus.trip_id)
#             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
#             .filter(TripStatus.time.between(from_date, to_date))
#             .distinct()
#             .all()
#         )

#         trip_ids = [t.trip_id for t in filtered_trips]

#         # ---- Pagination logic ----
#         total_items = len(trip_ids)
#         total_pages = (total_items + limit - 1) // limit
#         start_index = (page - 1) * limit
#         paginated_trip_ids = trip_ids[start_index:start_index+limit]

#         results = []

#         for trip_id in paginated_trip_ids:
#             document = Document.query.filter_by(Trip_id=trip_id).first()
#             if not document:
#                 continue

#             statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

#             # ---- gate entry / exit ----
#             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
#             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

#             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
#             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

#             trip_duration = None
#             if pravesh_entry_time and pravesh_exit_time:
#                 trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

#             # ---- dock in / out ----
#             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
#             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

#             dock_in_time = dock_in.time if dock_in else None
#             dock_out_time = dock_out.time if dock_out else None

#             docked_duration = None
#             if dock_in_time and dock_out_time:
#                 docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

#             # ---- dock details ----
#             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
#             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

#             dock_details_list = [{
#                 "docked_location": d.docked_location,
#                 "remarks": d.remarks,
#                 "user_dockin": d.user_dockin,
#                 "material_category": d.material_category,
#             } for d in dock_details]

#             # ---- driver name extraction ----
#             driver_name = None
#             if document.extracted_text_driver:
#                 try:
#                     extracted = document.extracted_text_driver if isinstance(document.extracted_text_driver, dict) else json.loads(document.extracted_text_driver or "{}")
#                     driver_name = extracted.get("driver_name") or extracted.get("name")
#                 except:
#                     driver_name = None

#             # ---- exit remark ----
#             exit_remark = pravesh_exit.remark if pravesh_exit else None

#             # ---- standard time lookup ----
#             status_value = "Green Channel" if document.green_channel else "Normal"
#             master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

#             standard_times = []
#             if master_record:
#                 for cat in material_categories:
#                     record = VehicleTypeDetails.query.filter_by(
#                         vehicle_type_id=master_record.id,
#                         status=status_value,
#                         category=cat
#                     ).first()
#                     if record and record.standard_time_minutes:
#                         standard_times.append(record.standard_time_minutes)

#             standard_time_duration = max(standard_times) if standard_times else None

#             # ---- efficiency ----
#             efficiency = None
#             efficiency_category = ""
#             if standard_time_duration and trip_duration and trip_duration > 0:
#                 efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}")
#                 efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"

#             # ---- response object ----
#             result = {
#                 "trip_id": trip_id,
#                 "vehicle_number": document.vehicle_number,
#                 "vehicle_type": document.vehicle_type,
#                 "transporter_name": document.transporter_name,
#                 "material_categories": material_categories,
#                 "loading_unloading": document.loading_or_unloading,
#                 "driver": driver_name,
#                 "vehicle_status": status_value,
#                 "bill_category": "",  # <-- Added empty
#                 "pravesh_entry_time": pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else None,
#                 "pravesh_exit_time": pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else None,
#                 "trip_duration": trip_duration,
#                 "dock_in_time": dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else None,
#                 "dock_out_time": dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else None,
#                 "docked_duration": docked_duration,
#                 "standard_time_duration": standard_time_duration,
#                 "extended_TAT": "",
#                 "efficiency": efficiency,
#                 "efficiency_category": efficiency_category,
#                 "dock_details": dock_details_list,
#                 "exit_remark": exit_remark,
#             }

#             results.append(result)

#         return jsonify({
#             "status": "success",
#             "page": page,
#             "limit": limit,
#             "total_items": total_items,
#             "total_pages": total_pages,
#             "data": results,
#         }), 200

#     except Exception as e:
#         print("Report Error:", str(e))
#         return jsonify({"status": "error", "message": str(e)}), 500



    


# # from openpyxl import Workbook
# # from openpyxl.styles import Font
# # import io
# # import json

# # @app.route("/download_trip_report", methods=["GET"])
# # def download_trip_report():
# #     try:
# #         from_date_str = request.args.get("from_date")
# #         to_date_str = request.args.get("to_date")

# #         if not from_date_str or not to_date_str:
# #             return jsonify({
# #                 "status": "error",
# #                 "message": "from_date and to_date query params are required in format YYYY-MM-DD HH:MM:SS"
# #             }), 400

# #         from_date = datetime.strptime(from_date_str, "%Y-%m-%d %H:%M:%S")
# #         to_date = datetime.strptime(to_date_str, "%Y-%m-%d %H:%M:%S")

# #         # 1️⃣ Get filtered trip_ids based on GATE ENTRY
# #         filtered_trips = (
# #             db.session.query(TripStatus.trip_id)
# #             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
# #             .filter(TripStatus.time.between(from_date, to_date))
# #             .distinct()
# #             .all()
# #         )

# #         trip_ids = [t.trip_id for t in filtered_trips]

# #         # 2️⃣ Prepare Excel workbook
# #         wb = Workbook()
# #         ws = wb.active
# #         ws.title = "Trip Report"

# #         headers = [
# #             "Trip ID",
# #             "Vehicle No",
# #             "Vehicle Type",
# #             "Transporter",
# #             "Material Categories",
# #             "Loading/Unloading",
# #             "Driver",
# #             "Green Channel",
# #             "Pravesh Entry Time",
# #             "Pravesh Exit Time",
# #             "Trip Duration (mins)",
# #             "Dock In Time",
# #             "Dock Out Time",
# #             "Docked Duration (mins)",
# #             "Standard Time Duration",
# #             "Extended TAT",
# #             "Efficiency",
# #             "Efficiency Category",
# #             "Docked Locations",
# #             "Dock User-DockIn",
# #             "Dock Remarks",
# #             "Exit Remark"
# #         ]
# #         ws.append(headers)
# #         for cell in ws[1]:
# #             cell.font = Font(bold=True)

# #         # 3️⃣ Build rows per trip
# #         for trip_id in trip_ids:
# #             document = Document.query.filter_by(Trip_id=trip_id).first()
# #             if not document:
# #                 continue

# #             statuses = (
# #                 TripStatus.query.filter_by(trip_id=trip_id)
# #                 .order_by(TripStatus.time.asc())
# #                 .all()
# #             )

# #             # Gate Entry / Exit
# #             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
# #             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

# #             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
# #             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

# #             trip_duration = None
# #             if pravesh_entry_time and pravesh_exit_time:
# #                 trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

# #             # Dock In / Out
# #             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
# #             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

# #             dock_in_time = dock_in.time if dock_in else None
# #             dock_out_time = dock_out.time if dock_out else None

# #             docked_duration = None
# #             if dock_in_time and dock_out_time:
# #                 docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

# #             # Dock Details & Materials
# #             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
# #             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

# #             docked_locations_set = {d.docked_location for d in dock_details if d.docked_location}
# #             dock_user_dockin_set = {d.user_dockin for d in dock_details if d.user_dockin}
# #             dock_remarks_set = {d.remarks for d in dock_details if d.remarks}

# #             docked_locations = ", ".join(sorted(docked_locations_set))
# #             dock_user_dockin = ", ".join(sorted(dock_user_dockin_set))
# #             dock_remarks = ", ".join(sorted(dock_remarks_set))

# #             # Safely extract driver name
# #             driver_name = None
# #             if document.extracted_text_driver:
# #                 try:
# #                     if isinstance(document.extracted_text_driver, dict):
# #                         extracted = document.extracted_text_driver
# #                     else:
# #                         extracted = json.loads(document.extracted_text_driver or "{}")
# #                     driver_name = extracted.get("driver_name") or extracted.get("name")
# #                 except Exception:
# #                     driver_name = None

# #             # 🔹 Exit remark from TripStatus (GATE EXIT)
# #             exit_remark = pravesh_exit.remark if pravesh_exit else ""

# #             ws.append([
# #                 trip_id,
# #                 document.vehicle_number or "",
# #                 document.vehicle_type or "",
# #                 document.transporter_name or "",
# #                 ", ".join(material_categories) if material_categories else "",
# #                 document.loading_or_unloading or "",
# #                 driver_name or "",
# #                 "Yes" if document.green_channel else "No",
# #                 pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
# #                 pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
# #                 trip_duration if trip_duration is not None else "",
# #                 dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
# #                 dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
# #                 docked_duration if docked_duration is not None else "",
# #                 "",   # standard_time_duration (future)
# #                 "",   # extended_TAT (future)
# #                 "",   # efficiency (future)
# #                 "",   # efficiency_category (future)
# #                 docked_locations,
# #                 dock_user_dockin,
# #                 dock_remarks,
# #                 exit_remark
# #             ])

# #         # 4️⃣ Send file as response
# #         output = io.BytesIO()
# #         wb.save(output)
# #         output.seek(0)

# #         filename = f"Trip_Report_{from_date_str.replace(':','-')}_to_{to_date_str.replace(':','-')}.xlsx"

# #         return send_file(
# #             output,
# #             as_attachment=True,
# #             download_name=filename,
# #             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
# #         )

# #     except Exception as e:
# #         print("Download Trip Report Error:", str(e))
# #         return jsonify({"status": "error", "message": str(e)}), 500


# from openpyxl import Workbook
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter

# import io
# import json
# from datetime import datetime
# from flask import send_file, jsonify, request

# @app.route("/download_trip_report", methods=["GET"])
# def download_trip_report():
#     try:
#         from_date_str = request.args.get("from_date")
#         to_date_str = request.args.get("to_date")

#         if not from_date_str or not to_date_str:
#             return jsonify({
#                 "status": "error",
#                 "message": "from_date and to_date required in format YYYY-MM-DD"
#             }), 400

#         from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
#         to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
#         from_date = datetime.combine(from_date, datetime.min.time())
#         to_date = datetime.combine(to_date, datetime.max.time())

#         filtered_trips = (
#             db.session.query(TripStatus.trip_id)
#             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
#             .filter(TripStatus.time.between(from_date, to_date))
#             .distinct()
#             .all()
#         )
#         trip_ids = [t.trip_id for t in filtered_trips]

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Trip Report"

#         headers = [
#             "Trip ID", "Vehicle No", "Vehicle Type", "Billable / Non Billable",
#             "Pravesh Entry Date", "Dock In Date", "Dock Out Date",
#             "Docked Duration", "Docked Duration (mins)",
#             "Pravesh Exit Date",
#             "Pravesh Trip Duration", "Pravesh Trip Duration (mins)",
#             "Standard Time Duration", "Extended TAT", "Efficiency",
#             "EFFICIENCY CATEGORY", "Transporter", "Material Category",
#             "Loading/Unloading", "Driver", "Dock User-DockIn",
#             "Docked Locations", "Dock Remarks", "Exit Remark"
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for trip_id in trip_ids:
#             document = Document.query.filter_by(Trip_id=trip_id).first()
#             if not document:
#                 continue

#             statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

#             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
#             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)
#             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
#             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

#             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
#             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)
#             dock_in_time = dock_in.time if dock_in else None
#             dock_out_time = dock_out.time if dock_out else None

#             trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60) if pravesh_entry_time and pravesh_exit_time else ""
#             docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60) if dock_in_time and dock_out_time else ""

#             # Duration Formatting
#             def format_duration(minutes):
#                 if minutes == "" or minutes is None:
#                     return "", ""
#                 hrs = minutes // 60
#                 mins = minutes % 60
#                 return f"{hrs}h{mins}m", minutes

#             docked_duration_formatted, docked_duration_minutes = format_duration(docked_duration)
#             trip_duration_formatted, trip_duration_minutes = format_duration(trip_duration)

#             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
#             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

#             status_value = "Green Channel" if document.green_channel else "Normal"
#             master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

#             standard_times = []
#             if master_record:
#                 for cat in material_categories:
#                     record = VehicleTypeDetails.query.filter_by(
#                         vehicle_type_id=master_record.id,
#                         status=status_value,
#                         category=cat
#                     ).first()
#                     if record and record.standard_time_minutes:
#                         standard_times.append(record.standard_time_minutes)

#             standard_time_duration = max(standard_times) if standard_times else ""
#             efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}") if standard_time_duration and trip_duration else ""
#             efficiency_category = "WITHIN" if efficiency and efficiency > 100 else ("EXCESS" if efficiency else "")

#             exit_remark = pravesh_exit.remark if pravesh_exit else ""

#             # Safe driver value extraction
#             driver_name = ""
#             if document.extracted_text_driver:
#                 try:
#                     extracted = (document.extracted_text_driver if isinstance(document.extracted_text_driver, dict)
#                                  else json.loads(document.extracted_text_driver))
#                     driver_name = extracted.get("driver_name") or extracted.get("name") or ""
#                 except:
#                     driver_name = ""

#             ws.append([
#                 trip_id,
#                 document.vehicle_number or "",
#                 document.vehicle_type or "",
#                 "",
#                 pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
#                 dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
#                 dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
#                 docked_duration_formatted,
#                 docked_duration_minutes,
#                 pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
#                 trip_duration_formatted,
#                 trip_duration_minutes,
#                 standard_time_duration,
#                 "",
#                 efficiency,
#                 efficiency_category,
#                 document.transporter_name or "",
#                 ", ".join(material_categories),
#                 document.loading_or_unloading or "",
#                 driver_name,
#                 ", ".join({d.user_dockin for d in dock_details if d.user_dockin}),
#                 ", ".join({d.docked_location for d in dock_details if d.docked_location}),
#                 ", ".join({d.remarks for d in dock_details if d.remarks}),
#                 exit_remark
#             ])

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         filename = f"Trip_Report_{from_date_str}_to_{to_date_str}.xlsx"
#         return send_file(output, as_attachment=True, download_name=filename,
#                          mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

#     except Exception as e:
#         print("Download Trip Report Error:", str(e))
#         return jsonify({"status": "error", "message": str(e)}), 500

# @app.route("/api/report/hrsdate-range", methods=["GET"])
# def hrs_report_date_range():
#     try:
#         # ---- GET query parameters ----
#         from_date_str = request.args.get("from_date")
#         to_date_str = request.args.get("to_date")
#         page = int(request.args.get("page", 1))
#         limit = int(request.args.get("limit", 50))

#         if not from_date_str or not to_date_str:
#             return jsonify({"status": "error", "message": "from_date & to_date required YYYY-MM-DD"}), 400

#         # ---- Parse only dates ----
#         from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
#         to_date = datetime.strptime(to_date_str, "%Y-%m-%d")

#         # ---- Convert to full datetime range ----
#         from_date = datetime.combine(from_date, datetime.min.time())   # 00:00:00
#         to_date = datetime.combine(to_date, datetime.max.time())       # 23:59:59

#         # ---- Filter TripStatus for GATE ENTRY ----
#         filtered_trips = (
#             db.session.query(TripStatus.trip_id)
#             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
#             .filter(TripStatus.time.between(from_date, to_date))
#             .distinct()
#             .all()
#         )

#         trip_ids = [t.trip_id for t in filtered_trips]

#         # ---- Pagination logic ----
#         total_items = len(trip_ids)
#         total_pages = (total_items + limit - 1) // limit
#         start_index = (page - 1) * limit
#         paginated_trip_ids = trip_ids[start_index:start_index+limit]

#         results = []

#         for trip_id in paginated_trip_ids:
#             document = Document.query.filter_by(Trip_id=trip_id).first()
#             if not document:
#                 continue

#             statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

#             # ---- gate entry / exit ----
#             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
#             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

#             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
#             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

#             trip_duration = None
#             if pravesh_entry_time and pravesh_exit_time:
#                 trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

#             # ---- dock in / out ----
#             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
#             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

#             dock_in_time = dock_in.time if dock_in else None
#             dock_out_time = dock_out.time if dock_out else None

#             docked_duration = None
#             if dock_in_time and dock_out_time:
#                 docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

#             # ---- HRS Band Calculation ----
#            # ---- HRS Band Calculation using PRAVESH ENTRY & EXIT ----
#             hrs_band = None
#             total_duration_hrs_band = None

#             if pravesh_entry_time and pravesh_exit_time:
#                 total_hours = (pravesh_exit_time - pravesh_entry_time).total_seconds() / 3600

#                 # HRS Band Rule (< 4.5 hours)
#                 hrs_band = "WITHIN" if total_hours < 4.5 else "EXCESS"

#                 # Format as Xh Ym
#                 hrs = int(total_hours)
#                 mins = int(round((total_hours - hrs) * 60))
#                 total_duration_hrs_band = f"{hrs}h {mins}m"


#             # ---- dock details ----
#             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
#             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

#             dock_details_list = [{
#                 "docked_location": d.docked_location,
#                 "remarks": d.remarks,
#                 "user_dockin": d.user_dockin,
#                 "material_category": d.material_category,
#             } for d in dock_details]

#             # ---- driver name extraction ----
#             driver_name = None
#             if document.extracted_text_driver:
#                 try:
#                     extracted = document.extracted_text_driver if isinstance(document.extracted_text_driver, dict) else json.loads(document.extracted_text_driver or "{}")
#                     driver_name = extracted.get("driver_name") or extracted.get("name")
#                 except:
#                     driver_name = None

#             # ---- exit remark ----
#             exit_remark = pravesh_exit.remark if pravesh_exit else None

#             # ---- standard time lookup ----
#             status_value = "Green Channel" if document.green_channel else "Normal"
#             master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

#             standard_times = []
#             if master_record:
#                 for cat in material_categories:
#                     record = VehicleTypeDetails.query.filter_by(
#                         vehicle_type_id=master_record.id,
#                         status=status_value,
#                         category=cat
#                     ).first()
#                     if record and record.standard_time_minutes:
#                         standard_times.append(record.standard_time_minutes)

#             standard_time_duration = max(standard_times) if standard_times else None

#             # ---- efficiency ----
#             efficiency = None
#             efficiency_category = ""
#             if standard_time_duration and trip_duration and trip_duration > 0:
#                 efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}")
#                 efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"

#             # ---- response object ----
#             result = {
#                 "trip_id": trip_id,
#                 "vehicle_number": document.vehicle_number,
#                 "vehicle_type": document.vehicle_type,
#                 "transporter_name": document.transporter_name,
#                 "material_categories": material_categories,
#                 "loading_unloading": document.loading_or_unloading,
#                 "driver": driver_name,
#                 "vehicle_status": status_value,
#                 "bill_category": "",
#                 "pravesh_entry_time": pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else None,
#                 "pravesh_exit_time": pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else None,
#                 "trip_duration": trip_duration,
#                 "dock_in_time": dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else None,
#                 "dock_out_time": dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else None,
#                 "docked_duration": docked_duration,
#                 "standard_time_duration": standard_time_duration,
#                 "extended_TAT": "",
#                 "efficiency": efficiency,
#                 "efficiency_category": efficiency_category,

#                 # ---- NEW FIELDS ADDED ----
#                 "hrs_band": hrs_band,
#                 "total_duration_hrs_band": total_duration_hrs_band,

#                 "dock_details": dock_details_list,
#                 "exit_remark": exit_remark,
#             }

#             results.append(result)

#         return jsonify({
#             "status": "success",
#             "page": page,
#             "limit": limit,
#             "total_items": total_items,
#             "total_pages": total_pages,
#             "data": results,
#         }), 200

#     except Exception as e:
#         print("Report Error:", str(e))
#         return jsonify({"status": "error", "message": str(e)}), 500

# @app.route("/download_hrs_report", methods=["GET"])
# def download_hrs_report():
#     try:
#         from_date_str = request.args.get("from_date")
#         to_date_str = request.args.get("to_date")

#         if not from_date_str or not to_date_str:
#             return jsonify({
#                 "status": "error",
#                 "message": "from_date and to_date required in format YYYY-MM-DD"
#             }), 400

#         from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
#         to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
#         from_date = datetime.combine(from_date, datetime.min.time())
#         to_date = datetime.combine(to_date, datetime.max.time())

#         filtered_trips = (
#             db.session.query(TripStatus.trip_id)
#             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
#             .filter(TripStatus.time.between(from_date, to_date))
#             .distinct()
#             .all()
#         )
#         trip_ids = [t.trip_id for t in filtered_trips]

#         wb = Workbook()
#         ws = wb.active
#         ws.title = "HrsBand Report"

#         headers = [
#             "Trip ID", "Vehicle No", "Vehicle Type", "Billable / Non Billable",
#             "Pravesh Entry Date", "Dock In Date", "Dock Out Date",
#             "Docked Duration", "Docked Duration (mins)",
#             "Pravesh Exit Date",
#             "Pravesh Trip Duration", "Pravesh Trip Duration (mins)",
#             "Standard Time Duration", "Extended TAT", "Efficiency",
#             "EFFICIENCY CATEGORY",

#             #### NEW FIELDS ####
#             "HRS Band", "Total Duration HRS Band",

#             "Transporter", "Material Category",
#             "Loading/Unloading", "Driver", "Dock User-DockIn",
#             "Docked Locations", "Dock Remarks", "Exit Remark"
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         for trip_id in trip_ids:
#             document = Document.query.filter_by(Trip_id=trip_id).first()
#             if not document:
#                 continue

#             statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

#             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
#             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)
#             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
#             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

#             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
#             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)
#             dock_in_time = dock_in.time if dock_in else None
#             dock_out_time = dock_out.time if dock_out else None

#             trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60) if pravesh_entry_time and pravesh_exit_time else ""
#             docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60) if dock_in_time and dock_out_time else ""

#             # Duration Formatting
#             def format_duration(minutes):
#                 if minutes == "" or minutes is None:
#                     return "", ""
#                 hrs = minutes // 60
#                 mins = minutes % 60
#                 return f"{hrs}h{mins}m", minutes

#             docked_duration_formatted, docked_duration_minutes = format_duration(docked_duration)
#             trip_duration_formatted, trip_duration_minutes = format_duration(trip_duration)

#             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
#             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

#             status_value = "Green Channel" if document.green_channel else "Normal"
#             master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

#             standard_times = []
#             if master_record:
#                 for cat in material_categories:
#                     record = VehicleTypeDetails.query.filter_by(
#                         vehicle_type_id=master_record.id,
#                         status=status_value,
#                         category=cat
#                     ).first()
#                     if record and record.standard_time_minutes:
#                         standard_times.append(record.standard_time_minutes)

#             standard_time_duration = max(standard_times) if standard_times else ""
#             efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}") if standard_time_duration and trip_duration else ""
#             efficiency_category = "WITHIN" if efficiency and efficiency > 100 else ("EXCESS" if efficiency else "")

#             exit_remark = pravesh_exit.remark if pravesh_exit else ""

#             # Safe driver extraction
#             driver_name = ""
#             if document.extracted_text_driver:
#                 try:
#                     extracted = (document.extracted_text_driver if isinstance(document.extracted_text_driver, dict)
#                                  else json.loads(document.extracted_text_driver))
#                     driver_name = extracted.get("driver_name") or extracted.get("name") or ""
#                 except:
#                     driver_name = ""

#             #### NEW → HRS BAND LOGIC ####
#             if trip_duration_minutes != "" and trip_duration_minutes is not None:
#                 hrs_band = "WITHIN" if trip_duration_minutes <= 270 else "EXCESS"
#                 total_duration_hrs_band = trip_duration_formatted
#             else:
#                 hrs_band = ""
#                 total_duration_hrs_band = ""

#             #### END NEW ####

#             ws.append([
#                 trip_id,
#                 document.vehicle_number or "",
#                 document.vehicle_type or "",
#                 "",
#                 pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
#                 dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
#                 dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
#                 docked_duration_formatted,
#                 docked_duration_minutes,
#                 pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
#                 trip_duration_formatted,
#                 trip_duration_minutes,
#                 standard_time_duration,
#                 "",
#                 efficiency,
#                 efficiency_category,

#                 #### NEW VALUES ####
#                 hrs_band,
#                 total_duration_hrs_band,

#                 document.transporter_name or "",
#                 ", ".join(material_categories),
#                 document.loading_or_unloading or "",
#                 driver_name,
#                 ", ".join({d.user_dockin for d in dock_details if d.user_dockin}),
#                 ", ".join({d.docked_location for d in dock_details if d.docked_location}),
#                 ", ".join({d.remarks for d in dock_details if d.remarks}),
#                 exit_remark
#             ])

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         filename = f"HrsBand_Report_{from_date_str}_to_{to_date_str}.xlsx"
#         return send_file(output, as_attachment=True, download_name=filename,
#                          mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

#     except Exception as e:
#         print("Download Trip Report Error:", str(e))
#         return jsonify({"status": "error", "message": str(e)}), 500

# @app.route("/api/report/dock-utilization", methods=["GET"])
# def dock_utilization_report():
#     try:
#         from_date = request.args.get("from_date")
#         to_date = request.args.get("to_date")

#         if not from_date or not to_date:
#             return jsonify({
#                 "status": "error",
#                 "message": "from_date & to_date required (YYYY-MM-DD)"
#             }), 400

#         start_dt = datetime.combine(
#             datetime.strptime(from_date, "%Y-%m-%d"),
#             datetime.min.time()
#         )
#         end_dt = datetime.combine(
#             datetime.strptime(to_date, "%Y-%m-%d"),
#             datetime.max.time()
#         )

#         statuses = (
#             db.session.query(TripStatus)
#             .filter(TripStatus.time.between(start_dt, end_dt))
#             .order_by(TripStatus.time.asc())
#             .all()
#         )

#         # -------- 24 Hour Buckets --------
#         report = {}
#         for h in range(24):
#             report[h] = {
#                 "hour": h,
#                 "slot": f"{datetime.strptime(f'{h:02d}', '%H').strftime('%I %p')} - "
#                         f"{datetime.strptime(f'{(h+1)%24:02d}', '%H').strftime('%I %p')}",

#                 # IN
#                 "in_vehicles": 0,
#                 "in_loading": 0,
#                 "in_unloading": 0,

#                 # OUT
#                 "out_vehicles": 0,
#                 "out_loading": 0,
#                 "out_unloading": 0
#             }

#         # -------- Process Records --------
#         for s in statuses:
#             hour = s.time.hour
#             location = (s.location or "").strip().upper()

#             doc = Document.query.filter_by(Trip_id=s.trip_id).first()
#             load_type = (doc.loading_or_unloading or "").upper()
#             normalized = load_type.replace("-", "").replace(" ", "")

#             is_loading = normalized.startswith("LOAD")
#             is_unloading = normalized.startswith("UNLOAD")

#             # -------- IN --------
#             if location in ["GATE ENTRY", "GATE ENTRYIN PROCESS", "PRAVESH IN"]:
#                 report[hour]["in_vehicles"] += 1

#                 if is_loading:
#                     report[hour]["in_loading"] += 1
#                 elif is_unloading:
#                     report[hour]["in_unloading"] += 1

#             # -------- OUT --------
#             elif location in ["GATE EXIT", "GATE OUTCLOSE", "PRAVESH OUT"]:
#                 report[hour]["out_vehicles"] += 1

#                 if is_loading:
#                     report[hour]["out_loading"] += 1
#                 elif is_unloading:
#                     report[hour]["out_unloading"] += 1

#         return jsonify({
#             "status": "success",
#             "from_date": from_date,
#             "to_date": to_date,
#             "data": list(report.values())
#         }), 200

#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({
#             "status": "error",
#             "message": str(e)
#         }), 500

# @app.route("/api/report/dock-utilization/download", methods=["GET"])
# def download_dock_utilization():
#     try:
#         from_date = request.args.get("from_date")
#         to_date = request.args.get("to_date")

#         if not from_date or not to_date:
#             return jsonify({"error": "from_date & to_date required"}), 400

#         start_dt = datetime.combine(datetime.strptime(from_date, "%Y-%m-%d"), datetime.min.time())
#         end_dt = datetime.combine(datetime.strptime(to_date, "%Y-%m-%d"), datetime.max.time())

#         statuses = (
#             db.session.query(TripStatus)
#             .filter(TripStatus.time.between(start_dt, end_dt))
#             .all()
#         )

#         # ---------- Workbook ----------
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Dock Utilization"

#         headers = [
#             "Hour Slot",
#             "IN Vehicles", "IN Loading", "IN Unloading",
#             "OUT Vehicles", "OUT Loading", "OUT Unloading"
#         ]
#         ws.append(headers)
#         for c in ws[1]:
#             c.font = Font(bold=True)

#         report = {h: {
#             "IN Vehicles": 0,
#             "IN Loading": 0,
#             "IN Unloading": 0,
#             "OUT Vehicles": 0,
#             "OUT Loading": 0,
#             "OUT Unloading": 0
#         } for h in range(24)}

#         for s in statuses:
#             hour = s.time.hour
#             location = (s.location or "").upper()

#             doc = Document.query.filter_by(Trip_id=s.trip_id).first()
#             load_type = (doc.loading_or_unloading or "").upper().replace(" ", "").replace("-", "")

#             is_loading = load_type.startswith("LOAD")
#             is_unloading = load_type.startswith("UNLOAD")

#             if location in ["GATE ENTRY", "GATE ENTRYIN PROCESS", "PRAVESH IN"]:
#                 report[hour]["IN Vehicles"] += 1
#                 if is_loading:
#                     report[hour]["IN Loading"] += 1
#                 elif is_unloading:
#                     report[hour]["IN Unloading"] += 1

#             elif location in ["GATE EXIT", "GATE OUTCLOSE", "PRAVESH OUT"]:
#                 report[hour]["OUT Vehicles"] += 1
#                 if is_loading:
#                     report[hour]["OUT Loading"] += 1
#                 elif is_unloading:
#                     report[hour]["OUT Unloading"] += 1

#         # ---------- Write rows ----------
#         for h in range(24):
#             slot = f"{datetime.strptime(f'{h:02d}', '%H').strftime('%I %p')} - " \
#                    f"{datetime.strptime(f'{(h+1)%24:02d}', '%H').strftime('%I %p')}"

#             ws.append([
#                 slot,
#                 report[h]["IN Vehicles"],
#                 report[h]["IN Loading"],
#                 report[h]["IN Unloading"],
#                 report[h]["OUT Vehicles"],
#                 report[h]["OUT Loading"],
#                 report[h]["OUT Unloading"],
#             ])

#         # ---------- TOTAL ROW ----------
#         ws.append([
#             "TOTAL",
#             sum(v["IN Vehicles"] for v in report.values()),
#             sum(v["IN Loading"] for v in report.values()),
#             sum(v["IN Unloading"] for v in report.values()),
#             sum(v["OUT Vehicles"] for v in report.values()),
#             sum(v["OUT Loading"] for v in report.values()),
#             sum(v["OUT Unloading"] for v in report.values()),
#         ])

#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=f"Dock_Utilization_{from_date}_to_{to_date}.xlsx",
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         import traceback
#         traceback.print_exc()
#         return jsonify({"error": str(e)}), 500



# # ------------------------------------------------------------
# # Run Flask App
# # ------------------------------------------------------------
# if __name__ == "__main__":
#     with app.app_context():
#         db.create_all()
#     app.run(
#         host="0.0.0.0",
#         port=5300,
#         debug=True,
#         ssl_context=("kirloskarWC2025.crt", "kirloskarWC2025.key")
#     )
from flask import Blueprint, Flask, jsonify, request
from flask_cors import CORS
from config import Config
from sqlalchemy import or_
from models import Document, db, VehicleTypeMaster, VehicleTypeDetails,DockInOutDetails,TripStatus, MasterVehicle, MasterDriver
from datetime import datetime
from sqlalchemy import text

app = Flask(__name__)
CORS(app)
app.config.from_object(Config)
db.init_app(app)



# ------------------------------------------------------------
# 1️⃣ Fetch only master table data
# ------------------------------------------------------------
@app.route("/get_vehicle_types", methods=["GET"])
def get_vehicle_types():
    """Returns only VehicleTypeMaster data"""
    try:
        masters = VehicleTypeMaster.query.all()
        result = [m.to_dict() for m in masters]
        return jsonify({
            "status": "success",
            "count": len(result),
            "data": result
        }), 200
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# ------------------------------------------------------------
# 2️⃣ Fetch master + details (joined)
# ------------------------------------------------------------
@app.route("/get_vehicle_type_details", methods=["GET"])
def get_vehicle_type_details():
    """Returns full master-detail vehicle type data"""
    try:
        results = (
            db.session.query(
                VehicleTypeMaster.id,
                VehicleTypeMaster.vehicle_type,
                VehicleTypeMaster.plant_location,
                VehicleTypeMaster.vehicle_type_created_by,
                VehicleTypeMaster.vehicle_type_created_time,
                VehicleTypeDetails.id.label("detail_id"),
                VehicleTypeDetails.status,
                VehicleTypeDetails.unload,
                VehicleTypeDetails.category,
                VehicleTypeDetails.standard_time_minutes,
                VehicleTypeDetails.standard_weight_kgs,
                VehicleTypeDetails.created_by,
                VehicleTypeDetails.created_time,
                VehicleTypeDetails.modified_by,
                VehicleTypeDetails.last_modified_time
            )
            .join(VehicleTypeDetails, VehicleTypeMaster.id == VehicleTypeDetails.vehicle_type_id)
            .order_by(VehicleTypeMaster.vehicle_type)
            .all()
        )

        data = [
            {
                "vehicle_type_master_id": r.id,
                "vehicle_type": r.vehicle_type,
                "plant_location": r.plant_location,
                "vehicle_type_created_by": r.vehicle_type_created_by,
                "vehicle_type_created_time": r.vehicle_type_created_time,
                "detail_id": r.detail_id,
                "status": r.status,
                "unload": r.unload,
                "category": r.category,
                "standard_time_minutes": r.standard_time_minutes,
                "standard_weight_kgs": r.standard_weight_kgs,
                "created_by": r.created_by,
                "created_time": r.created_time,
                "modified_by": r.modified_by,
                "last_modified_time": r.last_modified_time
            }
            for r in results
        ]

        return jsonify({
            "status": "success",
            "count": len(data),
            "data": data
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# ------------------------------------------------------------
# 3️⃣ Create new vehicle type (no duplicates)
# ------------------------------------------------------------
@app.route("/create_vehicle_type", methods=["POST"])
def create_vehicle_type():
    """Add new vehicle type — no duplicates allowed"""
    try:
        data = request.get_json()
        vehicle_type = data.get("vehicle_type", "").strip()
        plant_location = data.get("plant_location", "KAGAL")
        created_by = data.get("created_by", "admin")

        if not vehicle_type:
            return jsonify({"status": "error", "message": "Vehicle type is required"}), 400

        # Check duplicate (case-insensitive)
        existing = VehicleTypeMaster.query.filter(
            db.func.lower(VehicleTypeMaster.vehicle_type) == vehicle_type.lower()
        ).first()
        if existing:
            return jsonify({"status": "error", "message": "Vehicle type already exists"}), 409

        new_type = VehicleTypeMaster(
            vehicle_type=vehicle_type,
            plant_location=plant_location,
            vehicle_type_created_by=created_by,
            vehicle_type_created_time=datetime.now()
        )

        db.session.add(new_type)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Vehicle type created successfully"
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ------------------------------------------------------------
# 4️⃣ Save / Update vehicle type detail
# ------------------------------------------------------------
@app.route("/save_vehicle_type_details", methods=["POST"])
def save_vehicle_type_details():
    """Insert or update details per dropdown combination"""
    try:
        data = request.get_json()
        vehicle_type_id = data.get("vehicle_type_id")
        status = data.get("status")
        unload = data.get("unload")
        category = data.get("category")
        standard_time_minutes = data.get("standard_time_minutes")
        standard_weight_kgs = data.get("standard_weight_kgs")
        created_by = data.get("created_by", "admin")

        if not all([vehicle_type_id, status, unload, category]):
            return jsonify({"status": "error", "message": "All dropdowns are required"}), 400

        existing = VehicleTypeDetails.query.filter_by(
            vehicle_type_id=vehicle_type_id,
            status=status,
            unload=unload,
            category=category
        ).first()

        if existing:
            existing.standard_time_minutes = standard_time_minutes
            existing.standard_weight_kgs = standard_weight_kgs
            existing.modified_by = created_by
            existing.last_modified_time = datetime.now()
        else:
            new_detail = VehicleTypeDetails(
                vehicle_type_id=vehicle_type_id,
                status=status,
                unload=unload,
                category=category,
                standard_time_minutes=standard_time_minutes,
                standard_weight_kgs=standard_weight_kgs,
                created_by=created_by,
                created_time=datetime.now()
            )
            db.session.add(new_detail)

        db.session.commit()
        return jsonify({"status": "success", "message": "Saved successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    

@app.route("/vehicle_type/<int:vehicle_type_id>", methods=["DELETE"])
def delete_vehicle_type(vehicle_type_id):
    try:
        # Get optional query params for deleting specific detail
        status = request.args.get("status")
        unload = request.args.get("unload")
        category = request.args.get("category")

        # Delete specific detail if all params are provided
        if status and unload and category:
            detail = VehicleTypeDetails.query.filter_by(
                vehicle_type_id=vehicle_type_id,
                status=status,
                unload=unload,
                category=category
            ).first()
            if not detail:
                return jsonify({"message": "No matching vehicle type detail found"}), 404

            db.session.delete(detail)
            db.session.commit()
            return jsonify({"message": "Vehicle type detail deleted successfully"}), 200

        # Else delete master and all related details
        master = VehicleTypeMaster.query.filter_by(id=vehicle_type_id).first()
        if not master:
            return jsonify({"message": "Vehicle type master not found"}), 404

        db.session.delete(master)
        db.session.commit()
        return jsonify({"message": "Vehicle type and all related details deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"Error deleting vehicle type: {str(e)}"}), 500
    
from flask import Blueprint, Flask, jsonify, request
from flask_cors import CORS
from config import Config
from sqlalchemy import or_
from models import db, VehicleTypeMaster, VehicleTypeDetails, MasterVehicle, MasterDriver
from datetime import datetime

app = Flask(__name__)
CORS(app)
app.config.from_object(Config)
db.init_app(app)



# ------------------------------------------------------------
# 1️⃣ Fetch only master table data
# ------------------------------------------------------------
@app.route("/get_vehicle_types", methods=["GET"])
def get_vehicle_types():
    """Returns only VehicleTypeMaster data"""
    try:
        masters = VehicleTypeMaster.query.all()
        result = [m.to_dict() for m in masters]
        return jsonify({
            "status": "success",
            "count": len(result),
            "data": result
        }), 200
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# ------------------------------------------------------------
# 2️⃣ Fetch master + details (joined)
# ------------------------------------------------------------
@app.route("/get_vehicle_type_details", methods=["GET"])
def get_vehicle_type_details():
    """Returns full master-detail vehicle type data"""
    try:
        results = (
            db.session.query(
                VehicleTypeMaster.id,
                VehicleTypeMaster.vehicle_type,
                VehicleTypeMaster.plant_location,
                VehicleTypeMaster.vehicle_type_created_by,
                VehicleTypeMaster.vehicle_type_created_time,
                VehicleTypeDetails.id.label("detail_id"),
                VehicleTypeDetails.status,
                VehicleTypeDetails.unload,
                VehicleTypeDetails.category,
                VehicleTypeDetails.standard_time_minutes,
                VehicleTypeDetails.standard_weight_kgs,
                VehicleTypeDetails.created_by,
                VehicleTypeDetails.created_time,
                VehicleTypeDetails.modified_by,
                VehicleTypeDetails.last_modified_time
            )
            .join(VehicleTypeDetails, VehicleTypeMaster.id == VehicleTypeDetails.vehicle_type_id)
            .order_by(VehicleTypeMaster.vehicle_type)
            .all()
        )

        data = [
            {
                "vehicle_type_master_id": r.id,
                "vehicle_type": r.vehicle_type,
                "plant_location": r.plant_location,
                "vehicle_type_created_by": r.vehicle_type_created_by,
                "vehicle_type_created_time": r.vehicle_type_created_time,
                "detail_id": r.detail_id,
                "status": r.status,
                "unload": r.unload,
                "category": r.category,
                "standard_time_minutes": r.standard_time_minutes,
                "standard_weight_kgs": r.standard_weight_kgs,
                "created_by": r.created_by,
                "created_time": r.created_time,
                "modified_by": r.modified_by,
                "last_modified_time": r.last_modified_time
            }
            for r in results
        ]

        return jsonify({
            "status": "success",
            "count": len(data),
            "data": data
        }), 200

    except Exception as e:
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500


# ------------------------------------------------------------
# 3️⃣ Create new vehicle type (no duplicates)
# ------------------------------------------------------------
@app.route("/create_vehicle_type", methods=["POST"])
def create_vehicle_type():
    """Add new vehicle type — no duplicates allowed"""
    try:
        data = request.get_json()
        vehicle_type = data.get("vehicle_type", "").strip()
        plant_location = data.get("plant_location", "KAGAL")
        created_by = data.get("created_by", "admin")

        if not vehicle_type:
            return jsonify({"status": "error", "message": "Vehicle type is required"}), 400

        # Check duplicate (case-insensitive)
        existing = VehicleTypeMaster.query.filter(
            db.func.lower(VehicleTypeMaster.vehicle_type) == vehicle_type.lower()
        ).first()
        if existing:
            return jsonify({"status": "error", "message": "Vehicle type already exists"}), 409

        new_type = VehicleTypeMaster(
            vehicle_type=vehicle_type,
            plant_location=plant_location,
            vehicle_type_created_by=created_by,
            vehicle_type_created_time=datetime.now()
        )

        db.session.add(new_type)
        db.session.commit()

        return jsonify({
            "status": "success",
            "message": "Vehicle type created successfully"
        }), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500


# ------------------------------------------------------------
# 4️⃣ Save / Update vehicle type detail
# ------------------------------------------------------------
@app.route("/save_vehicle_type_details", methods=["POST"])
def save_vehicle_type_details():
    """Insert or update details per dropdown combination"""
    try:
        data = request.get_json()
        vehicle_type_id = data.get("vehicle_type_id")
        status = data.get("status")
        unload = data.get("unload")
        category = data.get("category")
        standard_time_minutes = data.get("standard_time_minutes")
        standard_weight_kgs = data.get("standard_weight_kgs")
        created_by = data.get("created_by", "admin")

        if not all([vehicle_type_id, status, unload, category]):
            return jsonify({"status": "error", "message": "All dropdowns are required"}), 400

        existing = VehicleTypeDetails.query.filter_by(
            vehicle_type_id=vehicle_type_id,
            status=status,
            unload=unload,
            category=category
        ).first()

        if existing:
            existing.standard_time_minutes = standard_time_minutes
            existing.standard_weight_kgs = standard_weight_kgs
            existing.modified_by = created_by
            existing.last_modified_time = datetime.now()
        else:
            new_detail = VehicleTypeDetails(
                vehicle_type_id=vehicle_type_id,
                status=status,
                unload=unload,
                category=category,
                standard_time_minutes=standard_time_minutes,
                standard_weight_kgs=standard_weight_kgs,
                created_by=created_by,
                created_time=datetime.now()
            )
            db.session.add(new_detail)

        db.session.commit()
        return jsonify({"status": "success", "message": "Saved successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500
    

@app.route("/vehicle_type/<int:vehicle_type_id>", methods=["DELETE"])
def delete_vehicle_type(vehicle_type_id):
    try:
        # Get optional query params for deleting specific detail
        status = request.args.get("status")
        unload = request.args.get("unload")
        category = request.args.get("category")

        # Delete specific detail if all params are provided
        if status and unload and category:
            detail = VehicleTypeDetails.query.filter_by(
                vehicle_type_id=vehicle_type_id,
                status=status,
                unload=unload,
                category=category
            ).first()
            if not detail:
                return jsonify({"message": "No matching vehicle type detail found"}), 404

            db.session.delete(detail)
            db.session.commit()
            return jsonify({"message": "Vehicle type detail deleted successfully"}), 200

        # Else delete master and all related details
        master = VehicleTypeMaster.query.filter_by(id=vehicle_type_id).first()
        if not master:
            return jsonify({"message": "Vehicle type master not found"}), 404

        db.session.delete(master)
        db.session.commit()
        return jsonify({"message": "Vehicle type and all related details deleted successfully"}), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"message": f"Error deleting vehicle type: {str(e)}"}), 500
    
vehicle_bp = Blueprint('vehicle_bp', __name__)

  
@app.route('/get_vehicles', methods=['GET'])
def get_vehicles():
    try:
        # --- 1️⃣ Get query params ---
        search = request.args.get('search', '', type=str)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)

        # --- 2️⃣ Base query ---
        query = MasterVehicle.query

        # --- 3️⃣ Apply search filter if provided ---
        if search:
            query = query.filter(MasterVehicle.vehicle_number.ilike(f"%{search}%"))

        # --- 4️⃣ Select only required fields AFTER filtering ---
        query = query.with_entities(MasterVehicle.id, MasterVehicle.vehicle_number)

        # --- 5️⃣ Pagination ---
        pagination = query.order_by(MasterVehicle.id.desc()).paginate(page=page, per_page=limit, error_out=False)

        # --- 6️⃣ Prepare response ---
        data = [{"id": v.id, "vehicle_number": v.vehicle_number} for v in pagination.items]

        return jsonify({
            "status": "success",
            "page": page,
            "total_pages": pagination.pages,
            "total_items": pagination.total,
            "vehicles": data
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

    

@app.route('/get_vehicles/<int:vehicle_id>', methods=['GET'])
def get_vehicle(vehicle_id):
    try:
        # Fetch vehicle details
        vehicle = MasterVehicle.query.get(vehicle_id)
        if not vehicle:
            return jsonify({'message': 'Vehicle not found'}), 404

        # Count number of times vehicle entered plant
        entry_count = Document.query.filter_by(vehicle_number=vehicle.vehicle_number).count()

        data = {
            'id': vehicle.id,
            'vehicle_number': vehicle.vehicle_number,
            'vehicle_type': vehicle.vehicle_type,
            'pravesh_remarks': vehicle.vehicle_doc_remark,
            'puc_validity': vehicle.puc_validity,
            'insurance_validity': vehicle.insurance_validity,
            'green_channel': vehicle.green_channel,
            'vehicle_banned': vehicle.vehicle_banned,
            'vehicle_banned_remarks': vehicle.vehicle_banned_remarks,
            'entry_count': entry_count  # 👈 new field added
        }

        return jsonify(data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
from sqlalchemy.exc import SQLAlchemyError


@app.route("/update_vehicle_status", methods=["PUT"])
def update_vehicle_status():
    try:
        data = request.get_json()
        vehicle_id = data.get("id")
        green_channel = data.get("green_channel")
        vehicle_banned = data.get("vehicle_banned")
        vehicle_banned_remarks = data.get("vehicle_banned_remarks")

        if not vehicle_id:
            return jsonify({"error": "Vehicle ID is required"}), 400

        vehicle = MasterVehicle.query.get(vehicle_id)
        if not vehicle:
            return jsonify({"error": "Vehicle not found"}), 404

        # ✅ Update only the provided fields
        if green_channel is not None:
            vehicle.green_channel = green_channel
        if vehicle_banned is not None:
            vehicle.vehicle_banned = vehicle_banned
        if vehicle_banned_remarks is not None:
            vehicle.vehicle_banned_remarks = vehicle_banned_remarks

        db.session.commit()

        return jsonify({
            "message": "Vehicle status updated successfully",
            "updated_data": {
                "id": vehicle.id,
                "green_channel": vehicle.green_channel,
                "vehicle_banned": vehicle.vehicle_banned,
                "vehicle_banned_remarks": vehicle.vehicle_banned_remarks
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500
    

@app.route('/get_drivers', methods=['GET'])
def fetch_drivers():
    try:
        search = request.args.get('search', '', type=str)
        page = request.args.get('page', 1, type=int)
        limit = request.args.get('limit', 20, type=int)

        query = MasterDriver.query

        if search:
            query = query.filter(
                db.or_(
                    MasterDriver.driver_license_number.ilike(f"%{search}%"),
                    MasterDriver.driver_mobile_number.ilike(f"%{search}%"),
                    MasterDriver.driver_vehicle_number.ilike(f"%{search}%")
                )
            )

        # Fetch extracted_text_driver also
        query = query.with_entities(
            MasterDriver.id,
            MasterDriver.driver_license_number,
            MasterDriver.driver_mobile_number,
            MasterDriver.driver_vehicle_number,
            MasterDriver.extracted_text_driver
        )

        pagination = query.order_by(MasterDriver.id.desc()).paginate(
            page=page, per_page=limit, error_out=False
        )

        data = []
        for d in pagination.items:
            # Parse extracted_text_driver → name
            try:
                extracted = (
                    d.extracted_text_driver
                    if isinstance(d.extracted_text_driver, dict)
                    else json.loads(d.extracted_text_driver or "{}")
                )
                name = extracted.get("name", "")
            except:
                name = ""

            data.append({
                "id": d.id,
                "driver_license_number": d.driver_license_number,
                "driver_mobile_number": d.driver_mobile_number,
                "driver_vehicle_number": d.driver_vehicle_number,
                "driver_name": name    # <-- added here
            })

        return jsonify({
            "status": "success",
            "page": page,
            "total_pages": pagination.pages,
            "total_items": pagination.total,
            "drivers": data
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({"status": "error", "message": str(e)}), 500



    

    


import json

@app.route('/get_driver/<int:driver_id>', methods=['GET'])
def fetch_driver_details(driver_id):
    try:
        driver = db.session.get(MasterDriver, driver_id)
        if not driver:
            return jsonify({'message': 'Driver not found'}), 404

        # Parse extracted text safely
        extracted_data = {}
        if driver.extracted_text_driver:
            try:
                extracted_data = (
                    driver.extracted_text_driver
                    if isinstance(driver.extracted_text_driver, dict)
                    else json.loads(driver.extracted_text_driver)
                )
            except:
                extracted_data = {}

        driver_name = extracted_data.get("name", "")

        trip_count = Document.query.filter_by(
            driver_license_number=driver.driver_license_number
        ).count()

        data = {
            'id': driver.id,
            'driver_mobile_number': driver.driver_mobile_number,
            'driver_license_number': driver.driver_license_number,
            'driver_vehicle_number': driver.driver_vehicle_number,
            'license_validity': driver.license_validity,
            'driver_name': driver_name,
            'created_at': driver.created_at,
            'driver_banned': driver.driver_banned,
            'driver_banned_remarks': driver.driver_banned_remarks,
            'trip_count': trip_count
        }

        return jsonify(data), 200

    except Exception as e:
        print("ERROR in get_driver:", e)
        return jsonify({'error': str(e)}), 500
    
    
    
@app.route("/update_driver_status", methods=["PUT"])
def update_driver_status():
    try:
        data = request.get_json()

        driver_id = data.get("id")
        driver_banned = data.get("driver_banned")
        driver_banned_remarks = data.get("driver_banned_remarks")

        if not driver_id:
            return jsonify({"error": "Driver ID is required"}), 400

        # ✅ SQLAlchemy 2.x safe way
        driver = db.session.get(MasterDriver, driver_id)
        if not driver:
            return jsonify({"error": "Driver not found"}), 404

        # ✅ Update only provided fields
        if driver_banned is not None:
            driver.driver_banned = driver_banned

        if driver_banned_remarks is not None:
            driver.driver_banned_remarks = driver_banned_remarks

        db.session.commit()

        return jsonify({
            "message": "Driver status updated successfully",
            "updated_data": {
                "id": driver.id,
                "driver_banned": driver.driver_banned,
                "driver_banned_remarks": driver.driver_banned_remarks
            }
        }), 200

    except Exception as e:
        db.session.rollback()
        print("Error in update_driver_status:", str(e))
        return jsonify({"error": str(e)}), 500


    

from datetime import datetime, timedelta
import pandas as pd
import os



@app.route('/get_report', methods=['GET'])
def get_report():
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')

        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

        documents = Document.query.filter(
            Document.timestamp >= start,
            Document.timestamp <= end
        ).all()

        results = []

        for doc in documents:

            # Fetch standard weight
            # vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
            # standard_weight = vt.standard_weight_kgs if vt else None
            # Fetch standard weight using vehicle_type_master

            standard_weight = ""
            
            result = db.session.execute(

                text("""

                    SELECT vtd.standard_weight_kgs

                    FROM vehicle_type_details vtd

                    JOIN vehicle_type_master vtm

                        ON vtm.id = vtd.vehicle_type_id

                    WHERE vtm.vehicle_type = :vehicle_type

                    LIMIT 1

                """),

                {"vehicle_type": doc.vehicle_type}

            ).fetchone()
            
            if result:

                standard_weight = result[0]

 

            # Fetch Dock-In & Dock-Out users
            dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
            user_dock_in = list({d.user_dockin for d in dock_details if d.user_dockin})
            user_dock_out = list({d.user_dockout for d in dock_details if d.user_dockout})

            # Handle "Other Transporter"
            transporter = doc.transporter_name
            if transporter == "Other":
                transporter = doc.other_transporter_name or transporter

            results.append({
                "Trip ID": doc.Trip_id,
                "Pravesh Entry Date": doc.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                "Vehicle No": doc.vehicle_number,
                "Total Invoices": doc.total_invoices,
                "Loading / Unloading": doc.loading_or_unloading,
                "Transporter": transporter,
                "Vehicle Type": doc.vehicle_type,
                "Standard Weight": standard_weight,
                "User-Dock-In": user_dock_in,
                "User-Dock-Out": user_dock_out
            })

        return jsonify({"status": "success", "data": results})

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/download_report', methods=['GET'])
def download_report():
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')

        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

        documents = Document.query.filter(
            Document.timestamp >= start,
            Document.timestamp <= end
        ).all()

        from openpyxl import Workbook
        from openpyxl.styles import Font
        import io

        wb = Workbook()
        ws = wb.active
        ws.title = "Report"

        headers = [
            "Trip ID",
            "Pravesh Entry Date",
            "Vehicle No",
            "Total Invoices",
            "Loading / Unloading",
            "Transporter",
            "Vehicle Type",
            "Standard Weight (kgs)",
            "User-Dock-In",
            "User-Dock-Out"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for doc in documents:
            # vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
            # standard_weight = vt.standard_weight_kgs if vt else None
            # Fetch standard weight using vehicle_type_master

            standard_weight = ""
            
            result = db.session.execute(

                text("""

                    SELECT vtd.standard_weight_kgs

                    FROM vehicle_type_details vtd

                    JOIN vehicle_type_master vtm

                        ON vtm.id = vtd.vehicle_type_id

                    WHERE vtm.vehicle_type = :vehicle_type

                    LIMIT 1

                """),

                {"vehicle_type": doc.vehicle_type}

            ).fetchone()
            
            if result:

                standard_weight = result[0]

 

            dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
            user_dock_in = ", ".join(list({d.user_dockin for d in dock_details if d.user_dockin}))
            user_dock_out = ", ".join(list({d.user_dockout for d in dock_details if d.user_dockout}))

            # Handle "Other Transporter"
            transporter = doc.transporter_name
            if transporter == "Other":
                transporter = doc.other_transporter_name or transporter

            ws.append([
                doc.Trip_id,
                doc.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                doc.vehicle_number,
                doc.total_invoices,
                doc.loading_or_unloading,
                transporter,
                doc.vehicle_type,
                standard_weight,
                user_dock_in,
                user_dock_out
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"Report_{start_date}_to_{end_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500
    


@app.route('/download_report_vehicle', methods=['GET'])
def download_report_vehicle():
    try:
        start_date = request.args.get('startDate')
        end_date = request.args.get('endDate')
        vehicle_number = request.args.get('vehicleNumber')

        if not vehicle_number or vehicle_number.strip() == "":
            return jsonify({"status": "error", "message": "vehicleNumber is required"}), 400

        vehicle_number = vehicle_number.strip()  # Remove spaces

        # Parse dates
        start = datetime.strptime(start_date, "%Y-%m-%d")
        end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)

        # Filter documents by date range and vehicle_number (case-insensitive, partial match)
        documents = Document.query.filter(
            Document.timestamp >= start,
            Document.timestamp < end,
            Document.vehicle_number.ilike(f"%{vehicle_number}%")
        ).all()

        if not documents:
            return jsonify({"status": "error", "message": "No records found for the given vehicle and date range"}), 404

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Report_{vehicle_number}"

        headers = [
            "Trip ID",
            "Pravesh Entry Date",
            "Vehicle No",
            "Total Invoices",
            "Loading / Unloading",
            "Transporter",
            "Vehicle Type",
            "Standard Weight (kgs)",
            "User-Dock-In",
            "User-Dock-Out"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for doc in documents:
            try:
                vt = VehicleTypeDetails.query.filter_by(category=doc.vehicle_type).first()
                standard_weight = vt.standard_weight_kgs if vt else None

                dock_details = DockInOutDetails.query.filter_by(trip_id=doc.Trip_id).all()
                user_dock_in = ", ".join(sorted({d.user_dockin for d in dock_details if d.user_dockin}))
                user_dock_out = ", ".join(sorted({d.user_dockout for d in dock_details if d.user_dockout}))

                transporter = doc.transporter_name
                if transporter == "Other":
                    transporter = doc.other_transporter_name or transporter

                ws.append([
                    doc.Trip_id or "",
                    doc.timestamp.strftime('%Y-%m-%d %H:%M:%S') if doc.timestamp else "",
                    doc.vehicle_number or "",
                    doc.total_invoices or 0,
                    doc.loading_or_unloading or "",
                    transporter,
                    doc.vehicle_type or "",
                    standard_weight or "",
                    user_dock_in,
                    user_dock_out
                ])
            except Exception as row_err:
                print(f"Skipping row due to error: {row_err}")
                continue

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"Report_{vehicle_number}_{start_date}_to_{end_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print(f"Download vehicle report error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500
    

from models import TripStatus, DockInOutDetails, Document  # ensure import

from models import TripStatus, DockInOutDetails, Document  # make sure this import exists
import json  # make sure this is imported
    
@app.route("/api/report/date-range", methods=["GET"])
def trip_report_date_range():
    try:
        # ---- GET query parameters ----
        from_date_str = request.args.get("from_date")
        to_date_str = request.args.get("to_date")
        page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 50))

        if not from_date_str or not to_date_str:
            return jsonify({"status": "error", "message": "from_date & to_date required YYYY-MM-DD"}), 400

        # ---- Parse only dates ----
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")

        # ---- Convert to full datetime range ----
        from_date = datetime.combine(from_date, datetime.min.time())   # 00:00:00
        to_date = datetime.combine(to_date, datetime.max.time())       # 23:59:59

        # ---- Filter TripStatus for GATE ENTRY ----
        filtered_trips = (
            db.session.query(TripStatus.trip_id)
            .filter(TripStatus.location.ilike("%GATE ENTRY%"))
            .filter(TripStatus.time.between(from_date, to_date))
            .distinct()
            .all()
        )

        trip_ids = [t.trip_id for t in filtered_trips]

        # ---- Pagination logic ----
        total_items = len(trip_ids)
        total_pages = (total_items + limit - 1) // limit
        start_index = (page - 1) * limit
        paginated_trip_ids = trip_ids[start_index:start_index+limit]

        results = []

        for trip_id in paginated_trip_ids:
            document = Document.query.filter_by(Trip_id=trip_id).first()
            if not document:
                continue

            statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

            # ---- gate entry / exit ----
            pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
            pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

            pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
            pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

            trip_duration = None
            if pravesh_entry_time and pravesh_exit_time:
                trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

            # ---- dock in / out ----
            dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
            dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

            dock_in_time = dock_in.time if dock_in else None
            dock_out_time = dock_out.time if dock_out else None

            docked_duration = None
            if dock_in_time and dock_out_time:
                docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

            # ---- dock details ----
            dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
            material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

            dock_details_list = [{
                "docked_location": d.docked_location,
                "remarks": d.remarks,
                "user_dockin": d.user_dockin,
                "material_category": d.material_category,
            } for d in dock_details]

            # ---- driver name extraction ----
            driver_name = None
            if document.extracted_text_driver:
                try:
                    extracted = document.extracted_text_driver if isinstance(document.extracted_text_driver, dict) else json.loads(document.extracted_text_driver or "{}")
                    driver_name = extracted.get("driver_name") or extracted.get("name")
                except:
                    driver_name = None

            # ---- exit remark ----
            exit_remark = pravesh_exit.remark if pravesh_exit else None

            # ---- standard time lookup ----
            # status_value = "Green Channel" if document.green_channel else "Normal"
            # master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

            # standard_times = []
            # if master_record:
            #     for cat in material_categories:
            #         record = VehicleTypeDetails.query.filter_by(
            #             vehicle_type_id=master_record.id,
            #             status=status_value,
            #             category=cat
            #         ).first()
            #         if record and record.standard_time_minutes:
            #             standard_times.append(record.standard_time_minutes)

            # standard_time_duration = max(standard_times) if standard_times else None
            status_value = "Green Channel" if document.green_channel else "Normal"
            master_record = VehicleTypeMaster.query.filter_by(
                vehicle_type=document.vehicle_type
            ).first()
            
            standard_time_duration = None
            
            # 1️⃣ CATEGORY-BASED STANDARD TIME (EXISTING LOGIC)
            standard_times = []
            if master_record and material_categories:
                for cat in material_categories:
                    record = VehicleTypeDetails.query.filter_by(
                        vehicle_type_id=master_record.id,
                        status=status_value,
                        category=cat
                    ).first()
            
                    if record and record.standard_time_minutes:
                        standard_times.append(record.standard_time_minutes)
            
            if standard_times:
                standard_time_duration = max(standard_times)
            
            # 2️⃣ FALLBACK: DEFAULT VEHICLE TYPE STANDARD TIME
            if not standard_time_duration and master_record:
                default_record = VehicleTypeDetails.query.filter_by(
                    vehicle_type_id=master_record.id,
                    category="DEFAULT"
                ).first()
            
                if default_record and default_record.standard_time_minutes:
                    standard_time_duration = default_record.standard_time_minutes

            # ---- efficiency ----
            efficiency = None
            efficiency_category = ""
            if standard_time_duration and trip_duration and trip_duration > 0:
                efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}")
                efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"

            # ---- response object ----
            result = {
                "trip_id": trip_id,
                "vehicle_number": document.vehicle_number,
                "vehicle_type": document.vehicle_type,
                "transporter_name": document.transporter_name,
                "material_categories": material_categories,
                "loading_unloading": document.loading_or_unloading,
                "driver": driver_name,
                "vehicle_status": status_value,
                "bill_category": "",  # <-- Added empty
                "pravesh_entry_time": pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else None,
                "pravesh_exit_time": pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else None,
                "trip_duration": trip_duration,
                "dock_in_time": dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else None,
                "dock_out_time": dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else None,
                "docked_duration": docked_duration,
                "standard_time_duration": standard_time_duration,
                "extended_TAT": "",
                "efficiency": efficiency,
                "efficiency_category": efficiency_category,
                "dock_details": dock_details_list,
                "exit_remark": exit_remark,
            }

            results.append(result)

        return jsonify({
            "status": "success",
            "page": page,
            "limit": limit,
            "total_items": total_items,
            "total_pages": total_pages,
            "data": results,
        }), 200

    except Exception as e:
        print("Report Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500



    


# from openpyxl import Workbook
# from openpyxl.styles import Font
# import io
# import json

# @app.route("/download_trip_report", methods=["GET"])
# def download_trip_report():
#     try:
#         from_date_str = request.args.get("from_date")
#         to_date_str = request.args.get("to_date")

#         if not from_date_str or not to_date_str:
#             return jsonify({
#                 "status": "error",
#                 "message": "from_date and to_date query params are required in format YYYY-MM-DD HH:MM:SS"
#             }), 400

#         from_date = datetime.strptime(from_date_str, "%Y-%m-%d %H:%M:%S")
#         to_date = datetime.strptime(to_date_str, "%Y-%m-%d %H:%M:%S")

#         # 1️⃣ Get filtered trip_ids based on GATE ENTRY
#         filtered_trips = (
#             db.session.query(TripStatus.trip_id)
#             .filter(TripStatus.location.ilike("%GATE ENTRY%"))
#             .filter(TripStatus.time.between(from_date, to_date))
#             .distinct()
#             .all()
#         )

#         trip_ids = [t.trip_id for t in filtered_trips]

#         # 2️⃣ Prepare Excel workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Trip Report"

#         headers = [
#             "Trip ID",
#             "Vehicle No",
#             "Vehicle Type",
#             "Transporter",
#             "Material Categories",
#             "Loading/Unloading",
#             "Driver",
#             "Green Channel",
#             "Pravesh Entry Time",
#             "Pravesh Exit Time",
#             "Trip Duration (mins)",
#             "Dock In Time",
#             "Dock Out Time",
#             "Docked Duration (mins)",
#             "Standard Time Duration",
#             "Extended TAT",
#             "Efficiency",
#             "Efficiency Category",
#             "Docked Locations",
#             "Dock User-DockIn",
#             "Dock Remarks",
#             "Exit Remark"
#         ]
#         ws.append(headers)
#         for cell in ws[1]:
#             cell.font = Font(bold=True)

#         # 3️⃣ Build rows per trip
#         for trip_id in trip_ids:
#             document = Document.query.filter_by(Trip_id=trip_id).first()
#             if not document:
#                 continue

#             statuses = (
#                 TripStatus.query.filter_by(trip_id=trip_id)
#                 .order_by(TripStatus.time.asc())
#                 .all()
#             )

#             # Gate Entry / Exit
#             pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
#             pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

#             pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
#             pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

#             trip_duration = None
#             if pravesh_entry_time and pravesh_exit_time:
#                 trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

#             # Dock In / Out
#             dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
#             dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

#             dock_in_time = dock_in.time if dock_in else None
#             dock_out_time = dock_out.time if dock_out else None

#             docked_duration = None
#             if dock_in_time and dock_out_time:
#                 docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

#             # Dock Details & Materials
#             dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
#             material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

#             docked_locations_set = {d.docked_location for d in dock_details if d.docked_location}
#             dock_user_dockin_set = {d.user_dockin for d in dock_details if d.user_dockin}
#             dock_remarks_set = {d.remarks for d in dock_details if d.remarks}

#             docked_locations = ", ".join(sorted(docked_locations_set))
#             dock_user_dockin = ", ".join(sorted(dock_user_dockin_set))
#             dock_remarks = ", ".join(sorted(dock_remarks_set))

#             # Safely extract driver name
#             driver_name = None
#             if document.extracted_text_driver:
#                 try:
#                     if isinstance(document.extracted_text_driver, dict):
#                         extracted = document.extracted_text_driver
#                     else:
#                         extracted = json.loads(document.extracted_text_driver or "{}")
#                     driver_name = extracted.get("driver_name") or extracted.get("name")
#                 except Exception:
#                     driver_name = None

#             # 🔹 Exit remark from TripStatus (GATE EXIT)
#             exit_remark = pravesh_exit.remark if pravesh_exit else ""

#             ws.append([
#                 trip_id,
#                 document.vehicle_number or "",
#                 document.vehicle_type or "",
#                 document.transporter_name or "",
#                 ", ".join(material_categories) if material_categories else "",
#                 document.loading_or_unloading or "",
#                 driver_name or "",
#                 "Yes" if document.green_channel else "No",
#                 pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
#                 pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
#                 trip_duration if trip_duration is not None else "",
#                 dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
#                 dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
#                 docked_duration if docked_duration is not None else "",
#                 "",   # standard_time_duration (future)
#                 "",   # extended_TAT (future)
#                 "",   # efficiency (future)
#                 "",   # efficiency_category (future)
#                 docked_locations,
#                 dock_user_dockin,
#                 dock_remarks,
#                 exit_remark
#             ])

#         # 4️⃣ Send file as response
#         output = io.BytesIO()
#         wb.save(output)
#         output.seek(0)

#         filename = f"Trip_Report_{from_date_str.replace(':','-')}_to_{to_date_str.replace(':','-')}.xlsx"

#         return send_file(
#             output,
#             as_attachment=True,
#             download_name=filename,
#             mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#         )

#     except Exception as e:
#         print("Download Trip Report Error:", str(e))
#         return jsonify({"status": "error", "message": str(e)}), 500


from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

import io
import json
from datetime import datetime
from flask import send_file, jsonify, request

@app.route("/download_trip_report", methods=["GET"])
def download_trip_report():
    try:
        from_date_str = request.args.get("from_date")
        to_date_str = request.args.get("to_date")

        if not from_date_str or not to_date_str:
            return jsonify({
                "status": "error",
                "message": "from_date and to_date required in format YYYY-MM-DD"
            }), 400

        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
        from_date = datetime.combine(from_date, datetime.min.time())
        to_date = datetime.combine(to_date, datetime.max.time())

        filtered_trips = (
            db.session.query(TripStatus.trip_id)
            .filter(TripStatus.location.ilike("%GATE ENTRY%"))
            .filter(TripStatus.time.between(from_date, to_date))
            .distinct()
            .all()
        )
        trip_ids = [t.trip_id for t in filtered_trips]

        wb = Workbook()
        ws = wb.active
        ws.title = "Trip Report"

        headers = [
            "Trip ID", "Vehicle No", "Vehicle Type", "Billable / Non Billable",
            "Pravesh Entry Date", "Dock In Date", "Dock Out Date",
            "Docked Duration", "Docked Duration (mins)",
            "Pravesh Exit Date",
            "Pravesh Trip Duration", "Pravesh Trip Duration (mins)",
            "Standard Time Duration", "Extended TAT", "Efficiency",
            "EFFICIENCY CATEGORY", "Transporter", "Material Category",
            "Loading/Unloading", "Driver", "Dock User-DockIn",
            "Docked Locations", "Dock Remarks", "Exit Remark"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for trip_id in trip_ids:
            document = Document.query.filter_by(Trip_id=trip_id).first()
            if not document:
                continue

            statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

            pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
            pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)
            pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
            pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

            dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
            dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)
            dock_in_time = dock_in.time if dock_in else None
            dock_out_time = dock_out.time if dock_out else None

            trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60) if pravesh_entry_time and pravesh_exit_time else ""
            docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60) if dock_in_time and dock_out_time else ""

            # Duration Formatting
            def format_duration(minutes):
                if minutes == "" or minutes is None:
                    return "", ""
                hrs = minutes // 60
                mins = minutes % 60
                return f"{hrs}h{mins}m", minutes

            docked_duration_formatted, docked_duration_minutes = format_duration(docked_duration)
            trip_duration_formatted, trip_duration_minutes = format_duration(trip_duration)

            dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
            material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

            status_value = "Green Channel" if document.green_channel else "Normal"
            master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

            # standard_times = []
            # if master_record:
            #     for cat in material_categories:
            #         record = VehicleTypeDetails.query.filter_by(
            #             vehicle_type_id=master_record.id,
            #             status=status_value,
            #             category=cat
            #         ).first()
            #         if record and record.standard_time_minutes:
            #             standard_times.append(record.standard_time_minutes)

            # standard_time_duration = max(standard_times) if standard_times else ""
            
            # ---- standard time lookup (same as frontend) ----
            standard_time_duration = None
            
            # 1️⃣ CATEGORY-BASED
            standard_times = []
            if master_record and material_categories:
                for cat in material_categories:
                    record = VehicleTypeDetails.query.filter_by(
                        vehicle_type_id=master_record.id,
                        status=status_value,
                        category=cat
                    ).first()
            
                    if record and record.standard_time_minutes:
                        standard_times.append(record.standard_time_minutes)
            
            if standard_times:
                standard_time_duration = max(standard_times)
            
            # 2️⃣ DEFAULT VEHICLE TYPE FALLBACK
            if not standard_time_duration and master_record:
                default_record = VehicleTypeDetails.query.filter_by(
                    vehicle_type_id=master_record.id,
                    category="DEFAULT"
                ).first()
            
                if default_record and default_record.standard_time_minutes:
                    standard_time_duration = default_record.standard_time_minutes
        
            efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}") if standard_time_duration and trip_duration else ""
            efficiency_category = "WITHIN" if efficiency and efficiency > 100 else ("EXCESS" if efficiency else "")

            exit_remark = pravesh_exit.remark if pravesh_exit else ""

            # Safe driver value extraction
            driver_name = ""
            if document.extracted_text_driver:
                try:
                    extracted = (document.extracted_text_driver if isinstance(document.extracted_text_driver, dict)
                                 else json.loads(document.extracted_text_driver))
                    driver_name = extracted.get("driver_name") or extracted.get("name") or ""
                except:
                    driver_name = ""

            ws.append([
                trip_id,
                document.vehicle_number or "",
                document.vehicle_type or "",
                "",
                pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
                dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
                dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
                docked_duration_formatted,
                docked_duration_minutes,
                pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
                trip_duration_formatted,
                trip_duration_minutes,
                standard_time_duration,
                "",
                efficiency,
                efficiency_category,
                document.transporter_name or "",
                ", ".join(material_categories),
                document.loading_or_unloading or "",
                driver_name,
                ", ".join({d.user_dockin for d in dock_details if d.user_dockin}),
                ", ".join({d.docked_location for d in dock_details if d.docked_location}),
                ", ".join({d.remarks for d in dock_details if d.remarks}),
                exit_remark
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Trip_Report_{from_date_str}_to_{to_date_str}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print("Download Trip Report Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/report/hrsdate-range", methods=["GET"])
def hrs_report_date_range():
    try:
        # ---- GET query parameters ----
        from_date_str = request.args.get("from_date")
        to_date_str = request.args.get("to_date")
        page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 50))

        if not from_date_str or not to_date_str:
            return jsonify({"status": "error", "message": "from_date & to_date required YYYY-MM-DD"}), 400

        # ---- Parse only dates ----
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")

        # ---- Convert to full datetime range ----
        from_date = datetime.combine(from_date, datetime.min.time())   # 00:00:00
        to_date = datetime.combine(to_date, datetime.max.time())       # 23:59:59

        # ---- Filter TripStatus for GATE ENTRY ----
        filtered_trips = (
            db.session.query(TripStatus.trip_id)
            .filter(TripStatus.location.ilike("%GATE ENTRY%"))
            .filter(TripStatus.time.between(from_date, to_date))
            .distinct()
            .all()
        )

        trip_ids = [t.trip_id for t in filtered_trips]

        # ---- Pagination logic ----
        total_items = len(trip_ids)
        total_pages = (total_items + limit - 1) // limit
        start_index = (page - 1) * limit
        paginated_trip_ids = trip_ids[start_index:start_index+limit]

        results = []

        for trip_id in paginated_trip_ids:
            document = Document.query.filter_by(Trip_id=trip_id).first()
            if not document:
                continue

            statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

            # ---- gate entry / exit ----
            pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
            pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

            pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
            pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

            trip_duration = None
            if pravesh_entry_time and pravesh_exit_time:
                trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

            # ---- dock in / out ----
            dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
            dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

            dock_in_time = dock_in.time if dock_in else None
            dock_out_time = dock_out.time if dock_out else None

            docked_duration = None
            if dock_in_time and dock_out_time:
                docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

            # ---- HRS Band Calculation ----
           # ---- HRS Band Calculation using PRAVESH ENTRY & EXIT ----
            hrs_band = None
            total_duration_hrs_band = None

            if pravesh_entry_time and pravesh_exit_time:
                total_hours = (pravesh_exit_time - pravesh_entry_time).total_seconds() / 3600

                # HRS Band Rule (< 4.5 hours)
                hrs_band = "WITHIN" if total_hours < 4.5 else "EXCESS"

                # Format as Xh Ym
                hrs = int(total_hours)
                mins = int(round((total_hours - hrs) * 60))
                total_duration_hrs_band = f"{hrs}h {mins}m"


            # ---- dock details ----
            dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
            material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

            dock_details_list = [{
                "docked_location": d.docked_location,
                "remarks": d.remarks,
                "user_dockin": d.user_dockin,
                "material_category": d.material_category,
            } for d in dock_details]

            # ---- driver name extraction ----
            driver_name = None
            if document.extracted_text_driver:
                try:
                    extracted = document.extracted_text_driver if isinstance(document.extracted_text_driver, dict) else json.loads(document.extracted_text_driver or "{}")
                    driver_name = extracted.get("driver_name") or extracted.get("name")
                except:
                    driver_name = None

            # ---- exit remark ----
            exit_remark = pravesh_exit.remark if pravesh_exit else None

            # # ---- standard time lookup ----
            # status_value = "Green Channel" if document.green_channel else "Normal"
            # master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

            # standard_times = []
            # if master_record:
            #     for cat in material_categories:
            #         record = VehicleTypeDetails.query.filter_by(
            #             vehicle_type_id=master_record.id,
            #             status=status_value,
            #             category=cat
            #         ).first()
            #         if record and record.standard_time_minutes:
            #             standard_times.append(record.standard_time_minutes)

            # standard_time_duration = max(standard_times) if standard_times else None
            
            # ---- standard time lookup ----
            status_value = "Green Channel" if document.green_channel else "Normal"
            master_record = VehicleTypeMaster.query.filter_by(
                vehicle_type=document.vehicle_type
            ).first()
            
            standard_time_duration = None
            
            # 1️⃣ CATEGORY-BASED STANDARD TIME (EXISTING LOGIC)
            standard_times = []
            if master_record and material_categories:
                for cat in material_categories:
                    record = VehicleTypeDetails.query.filter_by(
                        vehicle_type_id=master_record.id,
                        status=status_value,
                        category=cat
                    ).first()
            
                    if record and record.standard_time_minutes:
                        standard_times.append(record.standard_time_minutes)
            
            if standard_times:
                standard_time_duration = max(standard_times)
            
            # 2️⃣ FALLBACK: DEFAULT VEHICLE TYPE STANDARD TIME
            if not standard_time_duration and master_record:
                default_record = VehicleTypeDetails.query.filter_by(
                    vehicle_type_id=master_record.id,
                    category="DEFAULT"
                ).first()
            
                if default_record and default_record.standard_time_minutes:
                    standard_time_duration = default_record.standard_time_minutes
        

            # ---- efficiency ----
            efficiency = None
            efficiency_category = ""
            if standard_time_duration and trip_duration and trip_duration > 0:
                efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}")
                efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"

            # ---- response object ----
            result = {
                "trip_id": trip_id,
                "vehicle_number": document.vehicle_number,
                "vehicle_type": document.vehicle_type,
                "transporter_name": document.transporter_name,
                "material_categories": material_categories,
                "loading_unloading": document.loading_or_unloading,
                "driver": driver_name,
                "vehicle_status": status_value,
                "bill_category": "",
                "pravesh_entry_time": pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else None,
                "pravesh_exit_time": pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else None,
                "trip_duration": trip_duration,
                "dock_in_time": dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else None,
                "dock_out_time": dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else None,
                "docked_duration": docked_duration,
                "standard_time_duration": standard_time_duration,
                "extended_TAT": "",
                "efficiency": efficiency,
                "efficiency_category": efficiency_category,

                # ---- NEW FIELDS ADDED ----
                "hrs_band": hrs_band,
                "total_duration_hrs_band": total_duration_hrs_band,

                "dock_details": dock_details_list,
                "exit_remark": exit_remark,
            }

            results.append(result)

        return jsonify({
            "status": "success",
            "page": page,
            "limit": limit,
            "total_items": total_items,
            "total_pages": total_pages,
            "data": results,
        }), 200

    except Exception as e:
        print("Report Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/download_hrs_report", methods=["GET"])
def download_hrs_report():
    try:
        from_date_str = request.args.get("from_date")
        to_date_str = request.args.get("to_date")

        if not from_date_str or not to_date_str:
            return jsonify({
                "status": "error",
                "message": "from_date and to_date required in format YYYY-MM-DD"
            }), 400

        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")
        from_date = datetime.combine(from_date, datetime.min.time())
        to_date = datetime.combine(to_date, datetime.max.time())

        filtered_trips = (
            db.session.query(TripStatus.trip_id)
            .filter(TripStatus.location.ilike("%GATE ENTRY%"))
            .filter(TripStatus.time.between(from_date, to_date))
            .distinct()
            .all()
        )
        trip_ids = [t.trip_id for t in filtered_trips]

        wb = Workbook()
        ws = wb.active
        ws.title = "HrsBand Report"

        headers = [
            "Trip ID", "Vehicle No", "Vehicle Type", "Billable / Non Billable",
            "Pravesh Entry Date", "Dock In Date", "Dock Out Date",
            "Docked Duration", "Docked Duration (mins)",
            "Pravesh Exit Date",
            "Pravesh Trip Duration", "Pravesh Trip Duration (mins)",
            "Standard Time Duration", "Extended TAT", "Efficiency",
            "EFFICIENCY CATEGORY",

            #### NEW FIELDS ####
            "HRS Band", "Total Duration HRS Band",

            "Transporter", "Material Category",
            "Loading/Unloading", "Driver", "Dock User-DockIn",
            "Docked Locations", "Dock Remarks", "Exit Remark"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for trip_id in trip_ids:
            document = Document.query.filter_by(Trip_id=trip_id).first()
            if not document:
                continue

            statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

            pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
            pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)
            pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
            pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

            dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
            dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)
            dock_in_time = dock_in.time if dock_in else None
            dock_out_time = dock_out.time if dock_out else None

            trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60) if pravesh_entry_time and pravesh_exit_time else ""
            docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60) if dock_in_time and dock_out_time else ""

            # Duration Formatting
            def format_duration(minutes):
                if minutes == "" or minutes is None:
                    return "", ""
                hrs = minutes // 60
                mins = minutes % 60
                return f"{hrs}h{mins}m", minutes

            docked_duration_formatted, docked_duration_minutes = format_duration(docked_duration)
            trip_duration_formatted, trip_duration_minutes = format_duration(trip_duration)

            dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
            material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

            status_value = "Green Channel" if document.green_channel else "Normal"
            master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

            # standard_times = []
            # if master_record:
            #     for cat in material_categories:
            #         record = VehicleTypeDetails.query.filter_by(
            #             vehicle_type_id=master_record.id,
            #             status=status_value,
            #             category=cat
            #         ).first()
            #         if record and record.standard_time_minutes:
            #             standard_times.append(record.standard_time_minutes)

            # standard_time_duration = max(standard_times) if standard_times else ""
            
            # ---- standard time lookup (same as frontend) ----
            standard_time_duration = None
            
            # 1️⃣ CATEGORY-BASED
            standard_times = []
            if master_record and material_categories:
                for cat in material_categories:
                    record = VehicleTypeDetails.query.filter_by(
                        vehicle_type_id=master_record.id,
                        status=status_value,
                        category=cat
                    ).first()
            
                    if record and record.standard_time_minutes:
                        standard_times.append(record.standard_time_minutes)
            
            if standard_times:
                standard_time_duration = max(standard_times)
            
            # 2️⃣ DEFAULT VEHICLE TYPE FALLBACK
            if not standard_time_duration and master_record:
                default_record = VehicleTypeDetails.query.filter_by(
                    vehicle_type_id=master_record.id,
                    category="DEFAULT"
                ).first()
            
                if default_record and default_record.standard_time_minutes:
                    standard_time_duration = default_record.standard_time_minutes
        
            efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}") if standard_time_duration and trip_duration else ""
            efficiency_category = "WITHIN" if efficiency and efficiency > 100 else ("EXCESS" if efficiency else "")

            exit_remark = pravesh_exit.remark if pravesh_exit else ""

            # Safe driver extraction
            driver_name = ""
            if document.extracted_text_driver:
                try:
                    extracted = (document.extracted_text_driver if isinstance(document.extracted_text_driver, dict)
                                 else json.loads(document.extracted_text_driver))
                    driver_name = extracted.get("driver_name") or extracted.get("name") or ""
                except:
                    driver_name = ""

            #### NEW → HRS BAND LOGIC ####
            if trip_duration_minutes != "" and trip_duration_minutes is not None:
                hrs_band = "WITHIN" if trip_duration_minutes <= 270 else "EXCESS"
                total_duration_hrs_band = trip_duration_formatted
            else:
                hrs_band = ""
                total_duration_hrs_band = ""

            #### END NEW ####

            ws.append([
                trip_id,
                document.vehicle_number or "",
                document.vehicle_type or "",
                "",
                pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else "",
                dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else "",
                dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else "",
                docked_duration_formatted,
                docked_duration_minutes,
                pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else "",
                trip_duration_formatted,
                trip_duration_minutes,
                standard_time_duration,
                "",
                efficiency,
                efficiency_category,

                #### NEW VALUES ####
                hrs_band,
                total_duration_hrs_band,

                document.transporter_name or "",
                ", ".join(material_categories),
                document.loading_or_unloading or "",
                driver_name,
                ", ".join({d.user_dockin for d in dock_details if d.user_dockin}),
                ", ".join({d.docked_location for d in dock_details if d.docked_location}),
                ", ".join({d.remarks for d in dock_details if d.remarks}),
                exit_remark
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"HrsBand_Report_{from_date_str}_to_{to_date_str}.xlsx"
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        print("Download Trip Report Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route("/api/report/dock-utilization", methods=["GET"])
def dock_utilization_report():
    try:
        from_date = request.args.get("from_date")
        to_date = request.args.get("to_date")

        if not from_date or not to_date:
            return jsonify({
                "status": "error",
                "message": "from_date & to_date required (YYYY-MM-DD)"
            }), 400

        start_dt = datetime.combine(
            datetime.strptime(from_date, "%Y-%m-%d"),
            datetime.min.time()
        )
        end_dt = datetime.combine(
            datetime.strptime(to_date, "%Y-%m-%d"),
            datetime.max.time()
        )

        statuses = (
            db.session.query(TripStatus)
            .filter(TripStatus.time.between(start_dt, end_dt))
            .order_by(TripStatus.time.asc())
            .all()
        )

        # -------- 24 Hour Buckets --------
        report = {}
        for h in range(24):
            report[h] = {
                "hour": h,
                "slot": f"{datetime.strptime(f'{h:02d}', '%H').strftime('%I %p')} - "
                        f"{datetime.strptime(f'{(h+1)%24:02d}', '%H').strftime('%I %p')}",

                # IN
                "in_vehicles": 0,
                "in_loading": 0,
                "in_unloading": 0,

                # OUT
                "out_vehicles": 0,
                "out_loading": 0,
                "out_unloading": 0
            }

        # -------- Process Records --------
        for s in statuses:
            hour = s.time.hour
            location = (s.location or "").strip().upper()

            doc = Document.query.filter_by(Trip_id=s.trip_id).first()
            load_type = (doc.loading_or_unloading or "").upper()
            normalized = load_type.replace("-", "").replace(" ", "")

            is_loading = normalized.startswith("LOAD")
            is_unloading = normalized.startswith("UNLOAD")

            # -------- IN --------
            if location in ["GATE ENTRY", "GATE ENTRYIN PROCESS", "PRAVESH IN"]:
                report[hour]["in_vehicles"] += 1

                if is_loading:
                    report[hour]["in_loading"] += 1
                elif is_unloading:
                    report[hour]["in_unloading"] += 1

            # -------- OUT --------
            elif location in ["GATE EXIT", "GATE OUTCLOSE", "PRAVESH OUT"]:
                report[hour]["out_vehicles"] += 1

                if is_loading:
                    report[hour]["out_loading"] += 1
                elif is_unloading:
                    report[hour]["out_unloading"] += 1

        return jsonify({
            "status": "success",
            "from_date": from_date,
            "to_date": to_date,
            "data": list(report.values())
        }), 200

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({
            "status": "error",
            "message": str(e)
        }), 500

@app.route("/api/report/dock-utilization/download", methods=["GET"])
def download_dock_utilization():
    try:
        from_date = request.args.get("from_date")
        to_date = request.args.get("to_date")

        if not from_date or not to_date:
            return jsonify({"error": "from_date & to_date required"}), 400

        start_dt = datetime.combine(datetime.strptime(from_date, "%Y-%m-%d"), datetime.min.time())
        end_dt = datetime.combine(datetime.strptime(to_date, "%Y-%m-%d"), datetime.max.time())

        statuses = (
            db.session.query(TripStatus)
            .filter(TripStatus.time.between(start_dt, end_dt))
            .all()
        )

        # ---------- Workbook ----------
        wb = Workbook()
        ws = wb.active
        ws.title = "Dock Utilization"

        headers = [
            "Hour Slot",
            "IN Vehicles", "IN Loading", "IN Unloading",
            "OUT Vehicles", "OUT Loading", "OUT Unloading"
        ]
        ws.append(headers)
        for c in ws[1]:
            c.font = Font(bold=True)

        report = {h: {
            "IN Vehicles": 0,
            "IN Loading": 0,
            "IN Unloading": 0,
            "OUT Vehicles": 0,
            "OUT Loading": 0,
            "OUT Unloading": 0
        } for h in range(24)}

        for s in statuses:
            hour = s.time.hour
            location = (s.location or "").upper()

            doc = Document.query.filter_by(Trip_id=s.trip_id).first()
            load_type = (doc.loading_or_unloading or "").upper().replace(" ", "").replace("-", "")

            is_loading = load_type.startswith("LOAD")
            is_unloading = load_type.startswith("UNLOAD")

            if location in ["GATE ENTRY", "GATE ENTRYIN PROCESS", "PRAVESH IN"]:
                report[hour]["IN Vehicles"] += 1
                if is_loading:
                    report[hour]["IN Loading"] += 1
                elif is_unloading:
                    report[hour]["IN Unloading"] += 1

            elif location in ["GATE EXIT", "GATE OUTCLOSE", "PRAVESH OUT"]:
                report[hour]["OUT Vehicles"] += 1
                if is_loading:
                    report[hour]["OUT Loading"] += 1
                elif is_unloading:
                    report[hour]["OUT Unloading"] += 1

        # ---------- Write rows ----------
        for h in range(24):
            slot = f"{datetime.strptime(f'{h:02d}', '%H').strftime('%I %p')} - " \
                   f"{datetime.strptime(f'{(h+1)%24:02d}', '%H').strftime('%I %p')}"

            ws.append([
                slot,
                report[h]["IN Vehicles"],
                report[h]["IN Loading"],
                report[h]["IN Unloading"],
                report[h]["OUT Vehicles"],
                report[h]["OUT Loading"],
                report[h]["OUT Unloading"],
            ])

        # ---------- TOTAL ROW ----------
        ws.append([
            "TOTAL",
            sum(v["IN Vehicles"] for v in report.values()),
            sum(v["IN Loading"] for v in report.values()),
            sum(v["IN Unloading"] for v in report.values()),
            sum(v["OUT Vehicles"] for v in report.values()),
            sum(v["OUT Loading"] for v in report.values()),
            sum(v["OUT Unloading"] for v in report.values()),
        ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"Dock_Utilization_{from_date}_to_{to_date}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500



# ------------------------------------------------------------
# Run Flask App
# ------------------------------------------------------------
if __name__ == "__main__":
    with app.app_context():
        db.create_all()
    app.run(
        host="0.0.0.0",
        port=5300,
        debug=True,
        ssl_context=("kirloskarWC2025.crt", "kirloskarWC2025.key")
    )

    
@app.route("/api/report/date-range", methods=["GET"])
def trip_report_date_range():
    try:
        # ---- GET query parameters ----
        from_date_str = request.args.get("from_date")
        to_date_str = request.args.get("to_date")
        page = int(request.args.get("page", 1))
        limit = int(request.args.get("limit", 50))

        if not from_date_str or not to_date_str:
            return jsonify({"status": "error", "message": "from_date & to_date required YYYY-MM-DD"}), 400

        # ---- Parse only dates ----
        from_date = datetime.strptime(from_date_str, "%Y-%m-%d")
        to_date = datetime.strptime(to_date_str, "%Y-%m-%d")

        # ---- Convert to full datetime range ----
        from_date = datetime.combine(from_date, datetime.min.time())   # 00:00:00
        to_date = datetime.combine(to_date, datetime.max.time())       # 23:59:59

        # ---- Filter TripStatus for GATE ENTRY ----
        filtered_trips = (
            db.session.query(TripStatus.trip_id)
            .filter(TripStatus.location.ilike("%GATE ENTRY%"))
            .filter(TripStatus.time.between(from_date, to_date))
            .distinct()
            .all()
        )

        trip_ids = [t.trip_id for t in filtered_trips]

        # ---- Pagination logic ----
        total_items = len(trip_ids)
        total_pages = (total_items + limit - 1) // limit
        start_index = (page - 1) * limit
        paginated_trip_ids = trip_ids[start_index:start_index+limit]

        results = []

        for trip_id in paginated_trip_ids:
            document = Document.query.filter_by(Trip_id=trip_id).first()
            if not document:
                continue

            statuses = TripStatus.query.filter_by(trip_id=trip_id).order_by(TripStatus.time.asc()).all()

            # ---- gate entry / exit ----
            pravesh_entry = next((s for s in statuses if "GATE ENTRY" in s.location.upper()), None)
            pravesh_exit = next((s for s in statuses if "GATE EXIT" in s.location.upper()), None)

            pravesh_entry_time = pravesh_entry.time if pravesh_entry else None
            pravesh_exit_time = pravesh_exit.time if pravesh_exit else None

            trip_duration = None
            if pravesh_entry_time and pravesh_exit_time:
                trip_duration = int((pravesh_exit_time - pravesh_entry_time).total_seconds() / 60)

            # ---- dock in / out ----
            dock_in = next((s for s in statuses if "IN" in s.location.upper()), None)
            dock_out = next((s for s in statuses[::-1] if "OUT" in s.location.upper()), None)

            dock_in_time = dock_in.time if dock_in else None
            dock_out_time = dock_out.time if dock_out else None

            docked_duration = None
            if dock_in_time and dock_out_time:
                docked_duration = int((dock_out_time - dock_in_time).total_seconds() / 60)

            # ---- dock details ----
            dock_details = DockInOutDetails.query.filter_by(trip_id=trip_id).all()
            material_categories = list(set([d.material_category for d in dock_details if d.material_category]))

            dock_details_list = [{
                "docked_location": d.docked_location,
                "remarks": d.remarks,
                "user_dockin": d.user_dockin,
                "material_category": d.material_category,
            } for d in dock_details]

            # ---- driver name extraction ----
            driver_name = None
            if document.extracted_text_driver:
                try:
                    extracted = document.extracted_text_driver if isinstance(document.extracted_text_driver, dict) else json.loads(document.extracted_text_driver or "{}")
                    driver_name = extracted.get("driver_name") or extracted.get("name")
                except:
                    driver_name = None

            # ---- exit remark ----
            exit_remark = pravesh_exit.remark if pravesh_exit else None

            # ---- standard time lookup ----
            # status_value = "Green Channel" if document.green_channel else "Normal"
            # master_record = VehicleTypeMaster.query.filter_by(vehicle_type=document.vehicle_type).first()

            # standard_times = []
            # if master_record:
            #     for cat in material_categories:
            #         record = VehicleTypeDetails.query.filter_by(
            #             vehicle_type_id=master_record.id,
            #             status=status_value,
            #             category=cat
            #         ).first()
            #         if record and record.standard_time_minutes:
            #             standard_times.append(record.standard_time_minutes)

            # standard_time_duration = max(standard_times) if standard_times else None
            status_value = "Green Channel" if document.green_channel else "Normal"
            master_record = VehicleTypeMaster.query.filter_by(
                vehicle_type=document.vehicle_type
            ).first()     
            
            standard_time_duration = None
            
            # 1️⃣ CATEGORY-BASED STANDARD TIME (EXISTING LOGIC)
            standard_times = []
            if master_record and material_categories:
                for cat in material_categories:
                    record = VehicleTypeDetails.query.filter_by(
                        vehicle_type_id=master_record.id,
                        status=status_value,
                        category=cat
                    ).first()
            
                    if record and record.standard_time_minutes:
                        standard_times.append(record.standard_time_minutes)
            
            if standard_times:
                standard_time_duration = max(standard_times)
            
            # 2️⃣ FALLBACK: DEFAULT VEHICLE TYPE STANDARD TIME
            if not standard_time_duration and master_record:
                default_record = VehicleTypeDetails.query.filter_by(
                    vehicle_type_id=master_record.id,
                    category="DEFAULT"
                ).first()
            
                if default_record and default_record.standard_time_minutes:
                    standard_time_duration = default_record.standard_time_minutes

            # ---- efficiency ----
            # efficiency = None
            # efficiency_category = ""
            # if standard_time_duration and trip_duration and trip_duration > 0:
            #     efficiency = float(f"{(standard_time_duration / trip_duration) * 100:.2f}")
            #     efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"
            efficiency = None
            efficiency_category = ""

            trip_duration_val = int(trip_duration) if str(trip_duration).isdigit() else 0
            standard_time_val = int(standard_time_duration) if str(standard_time_duration).isdigit() else 0

            if trip_duration_val == 0:
                efficiency = 100
                efficiency_category = "WITHIN"
            else:
                efficiency = round((trip_duration_val / standard_time_val ) * 100, 2)
                efficiency_category = "WITHIN" if efficiency > 100 else "EXCESS"



            # ---- response object ----
            result = {
                "trip_id": trip_id,
                "vehicle_number": document.vehicle_number,
                "vehicle_type": document.vehicle_type,
                "transporter_name": document.transporter_name,
                "material_categories": material_categories,
                "loading_unloading": document.loading_or_unloading,
                "driver": driver_name,
                "vehicle_status": status_value,
                "bill_category": "",  # <-- Added empty
                "pravesh_entry_time": pravesh_entry_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_entry_time else None,
                "pravesh_exit_time": pravesh_exit_time.strftime("%Y-%m-%d %H:%M:%S") if pravesh_exit_time else None,
                "trip_duration": trip_duration,
                "dock_in_time": dock_in_time.strftime("%Y-%m-%d %H:%M:%S") if dock_in_time else None,
                "dock_out_time": dock_out_time.strftime("%Y-%m-%d %H:%M:%S") if dock_out_time else None,
                "docked_duration": docked_duration,
                "standard_time_duration": standard_time_duration,
                "extended_TAT": "",
                "efficiency": efficiency,
                "efficiency_category": efficiency_category,
                "dock_details": dock_details_list,
                "exit_remark": exit_remark,
            }

            results.append(result)

        return jsonify({
            "status": "success",
            "page": page,
            "limit": limit,
            "total_items": total_items,
            "total_pages": total_pages,
            "data": results,
        }), 200

    except Exception as e:
        print("Report Error:", str(e))
        return jsonify({"status": "error", "message": str(e)}), 500

